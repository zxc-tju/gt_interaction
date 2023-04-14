"""
Interaction simulator
"""
import copy
import scipy.io
import math
import os.path
from multiprocessing import Process, Manager
import numpy as np
import pandas as pd
from agent import Agent
from tools.utility import draw_rectangle, get_central_vertices
from tools.lattice_planner import lattice_planning
import matplotlib.pyplot as plt
from NDS_analysis import analyze_ipv_in_nds, cal_pet
import xlsxwriter
import time
from tqdm import tqdm
from time import gmtime, strftime
import seaborn as sns
from openpyxl import load_workbook

sigma = 1.5
INTERACTION_DIS = 4


class Scenario:
    def __init__(self, pos, vel, heading, ipv, conl_type, sce_type='left turn'):

        self.position = [np.array(pos[0]), np.array(pos[1])]
        self.velocity = [np.array(vel[0]), np.array(vel[1])]

        self.heading = [np.array(heading[0]), np.array(heading[1])]
        self.ipv = [np.array(ipv[0]), np.array(ipv[1])]
        self.conl_type = [conl_type[0], conl_type[1]]
        if sce_type == 'left turn':
            self.target = ['lt', 'gs']
        elif sce_type == 'ramp merge':
            self.target = ['ml', 'ir']


class Simulator:
    def __init__(self, case_id=None):
        self.gs_id = 0
        self.sim_type = None
        self.semantic_result = None
        self.output_directory = None
        self.tag = None
        self.case_id = case_id
        self.scenario = None
        self.agent_lt = None
        self.agent_gs = None
        self.num_step = 0
        self.simu_time = 0
        self.case_len = 0
        self.ending_point = None
        self.gs_actual_trj = []
        self.lt_actual_trj = []
        self.ipv_list = []

    def initialize(self, scenario, case_tag):
        self.scenario = scenario

        self.agent_lt = Agent(scenario.position[0], scenario.velocity[0], scenario.heading[0], scenario.target[0], )
        self.agent_gs = Agent(scenario.position[1], scenario.velocity[1], scenario.heading[1], scenario.target[1], )
        self.agent_lt.estimated_inter_agent = [copy.deepcopy(self.agent_gs)]
        self.agent_gs.estimated_inter_agent = [copy.deepcopy(self.agent_lt)]
        self.agent_lt.ipv = self.scenario.ipv[0]
        self.agent_gs.ipv = self.scenario.ipv[1]
        self.agent_lt.conl_type = self.scenario.conl_type[0]
        self.agent_gs.conl_type = self.scenario.conl_type[1]
        self.tag = case_tag

    def interact(self, simu_step=30, iter_limit=30,
                 make_video=False,
                 break_when_finish=False,
                 interactive=True,
                 file_path='/'
                 ):
        """
        Simulate the given scenario step by step

        Parameters
        ----------
        file_path
        interactive
        iter_limit: max iteration number
        make_video
        simu_step: number of simulation steps
        break_when_finish: (if set to be True) break the simulation when any agent crossed the conflict point

        """
        self.num_step = simu_step
        # iter_limit = 3

        if make_video:
            plt.ion()
            _, ax = plt.subplots(figsize=[8, 8])

        for t in range(self.num_step):

            # print('time_step: ', t, '/', self.num_step)

            "==plan for left-turn=="
            if self.agent_lt.conl_type in {'linear-gt'}:
                # time1 = time.perf_counter()
                self.agent_lt.lp_ibr_interact(iter_limit=iter_limit, interactive=interactive)
                # time2 = time.perf_counter()
                # print('time consumption: ', time2 - time1)

            elif self.agent_lt.conl_type in {'gt', 'opt'}:

                # ==interaction with parallel virtual agents
                # self.agent_lt.ibr_interact_with_virtual_agents(self.agent_gs)
                # self.agent_lt.ibr_interact_with_virtual_agents_parallel(self.agent_gs)
                iter_limit_lt = iter_limit
                if self.agent_lt.conl_type == 'opt':
                    iter_limit_lt = 0
                # ==interaction with estimated agent
                self.agent_lt.ibr_interact(iter_limit=iter_limit_lt)

            elif self.agent_lt.conl_type in {'idm'}:
                self.agent_lt.idm_plan(self.agent_gs)

            elif self.agent_lt.conl_type in {'replay'}:
                t_end = t + self.agent_lt.track_len
                self.agent_lt.trj_solution = self.lt_actual_trj[t:t_end, :]

            elif self.agent_lt.conl_type in {'lattice'}:
                role = 'lt_nds'
                if self.sim_type == 'simu_left_turn':
                    role = 'lt'
                elif self.sim_type == 'simu_ramp':
                    role = 'ml'
                if not self.agent_lt.plan_count:
                    path_point, _ = get_central_vertices(role,
                                                         origin_point=self.agent_lt.observed_trajectory[0, 0:2])
                    obstacle_data = {'px': self.agent_gs.position[0],
                                     'py': self.agent_gs.position[1],
                                     'v': np.linalg.norm(self.agent_gs.velocity),
                                     'heading': self.agent_gs.heading}
                    initial_state = {'px': self.agent_lt.position[0],
                                     'py': self.agent_lt.position[1],
                                     'v': np.linalg.norm(self.agent_lt.velocity),
                                     'heading': self.agent_lt.heading}
                    res, state = lattice_planning(path_point, obstacle_data, initial_state, show_res=False)
                    if len(res) >= 3:
                        self.agent_lt.trj_solution = np.array(res[:self.agent_lt.track_len])
                        self.agent_lt.plan_count = 2
                        if state == 'planning_back':
                            self.agent_lt.plan_count = 7
                    else:
                        self.agent_lt.cruise_plan()

                else:
                    self.agent_lt.trj_solution = self.agent_lt.trj_solution[1:, :]
                    self.agent_lt.plan_count = self.agent_lt.plan_count - 1

            "==plan for go straight=="
            if self.agent_gs.conl_type in {'linear-gt'}:
                self.agent_gs.lp_ibr_interact(iter_limit=iter_limit, interactive=interactive)

            elif self.agent_gs.conl_type in {'gt', 'opt'}:
                # ==interaction with parallel virtual agents
                # self.agent_gs.ibr_interact_with_virtual_agents(self.agent_lt, iter_limit)
                # self.agent_gs.ibr_interact_with_virtual_agents_parallel(self.agent_lt, iter_limit)

                iter_limit_gs = iter_limit
                if self.agent_gs.conl_type == 'opt':
                    iter_limit_gs = 0
                # ==interaction with estimated agent
                self.agent_gs.ibr_interact(iter_limit=iter_limit_gs)

            elif self.agent_gs.conl_type in {'idm'}:
                self.agent_gs.idm_plan(self.agent_lt)

            elif self.agent_gs.conl_type in {'replay'}:
                track_len = self.agent_gs.track_len
                t_end = t + track_len
                self.agent_gs.trj_solution = self.gs_actual_trj[t:t_end, :]

            elif self.agent_gs.conl_type in {'lattice'}:
                role = 'gs_nds'
                if self.sim_type == 'simu_left_turn':
                    role = 'gs'
                elif self.sim_type == 'simu_ramp':
                    role = 'ir'
                if not self.agent_gs.plan_count:
                    path_point, _ = get_central_vertices(role, origin_point=self.agent_gs.position)
                    obstacle_data = {'px': self.agent_lt.position[0],
                                     'py': self.agent_lt.position[1],
                                     'v': np.linalg.norm(self.agent_lt.velocity),
                                     'heading': self.agent_lt.heading}
                    initial_state = {'px': self.agent_gs.position[0],
                                     'py': self.agent_gs.position[1],
                                     'v': np.linalg.norm(self.agent_gs.velocity),
                                     'heading': self.agent_gs.heading}
                    res, state = lattice_planning(path_point, obstacle_data, initial_state, show_res=False)
                    if len(res) >= 3:
                        self.agent_gs.trj_solution = np.array(res[:self.agent_gs.track_len])
                        self.agent_gs.plan_count = 2
                        if state == 'planning_back':
                            self.agent_gs.plan_count = 7
                    else:
                        self.agent_gs.cruise_plan()

                else:
                    self.agent_gs.trj_solution = self.agent_gs.trj_solution[1:, :]
                    self.agent_gs.plan_count = self.agent_gs.plan_count - 1

            "==update state=="
            self.agent_lt.update_state([self.agent_gs])
            self.agent_gs.update_state([self.agent_lt])

            "==update video=="
            if make_video:
                plt.cla()
                if self.sim_type == 'simu_left_turn':
                    img = plt.imread('background_pic/T_intersection.jpg')
                    plt.imshow(img, extent=[-9.1, 24.9, -13, 8])
                    plt.xlim([-9.1, 24.9])
                    plt.ylim([-13, 8])
                    # central vertices
                    cv_it, _ = get_central_vertices('lt')
                    cv_gs, _ = get_central_vertices('gs')
                elif self.sim_type == 'simu_ramp':
                    cv_it, _ = get_central_vertices('ml')
                    cv_gs, _ = get_central_vertices('ir')
                    plt.xlim([-40, 20])
                    plt.ylim([-15, 10])
                    img = plt.imread('background_pic/ramp.png')
                    plt.imshow(img, extent=[-40, 20, -15, 10])
                elif self.sim_type == 'nds':
                    plt.xlim([-22 - 13, 53 - 13])
                    plt.ylim([-31 - 7.8, 57 - 7.8])
                    img = plt.imread('background_pic/Jianhexianxia-v2.png')
                    plt.imshow(img, extent=[-28 - 13, 58 - 13, -42 - 7.8, 64 - 7.8])
                    # central vertices
                    lt_origin_point = self.agent_lt.observed_trajectory[0, 0:2]
                    gs_origin_point = self.agent_gs.observed_trajectory[0, 0:2]
                    cv_it, _ = get_central_vertices('lt_nds', origin_point=lt_origin_point)
                    cv_gs, _ = get_central_vertices('gs_nds', origin_point=gs_origin_point)

                draw_rectangle(self.agent_lt.position[0], self.agent_lt.position[1],
                               self.agent_lt.heading / math.pi * 180, ax,
                               para_alpha=1, para_color='#0E76CF')
                draw_rectangle(self.agent_gs.position[0], self.agent_gs.position[1],
                               self.agent_gs.heading / math.pi * 180, ax,
                               para_alpha=1, para_color='#7030A0')

                plt.plot(self.agent_lt.trj_solution[:10, 0], self.agent_lt.trj_solution[:10, 1],
                         color='blue', alpha=0.5)
                plt.plot(self.agent_gs.trj_solution[:10, 0], self.agent_gs.trj_solution[:10, 1],
                         color='red', alpha=0.5)

                ax.axis('scaled')
                plt.show()
                plt.pause(0.0001)
                plt.savefig(file_path + self.tag + '-' + str(t) + '.png',
                            dpi=300)

            if break_when_finish:
                if self.agent_gs.observed_trajectory[-1, 0] < self.agent_lt.observed_trajectory[-1, 0] \
                        or self.agent_lt.observed_trajectory[-1, 1] > self.agent_gs.observed_trajectory[-1, 1]:
                    self.num_step = t + 1
                    break

    def visualize_final_results(self, file_path):

        cv_1 = []
        cv_2 = []
        # set figures
        fig, axes = plt.subplots(1, 2, figsize=[10, 5])
        # fig.suptitle('trajectory_LT_' + self.semantic_result)
        axes[0].set_title('trajectory')
        axes[1].set_title('velocity')
        if self.sim_type == 'simu_left_turn':
            axes[0].set(aspect=1, xlim=(-9.1, 24.9), ylim=(-13, 8))
            img = plt.imread('background_pic/T_intersection.jpg')
            axes[0].imshow(img, extent=[-9.1, 24.9, -13, 8])
            # central vertices
            cv_1, _ = get_central_vertices('lt')
            cv_2, _ = get_central_vertices('gs')
        elif self.sim_type == 'simu_ramp':
            cv_1, _ = get_central_vertices('ml')
            cv_2, _ = get_central_vertices('ir')
            axes[0].set(aspect=1, xlim=(-40, 20), ylim=(-15, 10))
            img = plt.imread('background_pic/ramp.png')
            axes[0].imshow(img, extent=[-40, 20, -15, 10])
        elif self.sim_type == 'nds':
            axes[0].set(aspect=1, xlim=(-22 - 13, 53 - 13), ylim=(-31 - 7.8, 57 - 7.8))
            img = plt.imread('background_pic/Jianhexianxia-v2.png')
            axes[0].imshow(img, extent=[-28 - 13, 58 - 13, -42 - 7.8, 64 - 7.8])
            # central vertices
            lt_origin_point = self.agent_lt.observed_trajectory[0, 0:2]
            gs_origin_point = self.agent_gs.observed_trajectory[0, 0:2]
            cv_1, _ = get_central_vertices('lt_nds', origin_point=lt_origin_point)
            cv_2, _ = get_central_vertices('gs_nds', origin_point=gs_origin_point)

        "---- data abstraction ----"
        # lt track (observed in simulation and ground truth in nds)
        lt_ob_trj = self.agent_lt.observed_trajectory[:, 0:2]
        lt_ob_heading = self.agent_lt.observed_trajectory[:, 4] / math.pi * 180

        vel_ob_vel_norm_lt = np.linalg.norm(self.agent_lt.observed_trajectory[:, 2:4], axis=1)
        if self.sim_type == 'nds':
            vel_nds_vel_norm_lt = np.linalg.norm(self.lt_actual_trj[:, 2:4], axis=1)
            lt_nds_trj = np.array(self.lt_actual_trj[:, 0:2])
            lt_nds_heading = np.array(self.lt_actual_trj[:, 4]) / math.pi * 180

        # gs track (observed in simulation and ground truth in nds)
        gs_ob_trj = self.agent_gs.observed_trajectory[:, 0:2]
        gs_ob_heading = self.agent_gs.observed_trajectory[:, 4] / math.pi * 180

        vel_ob_vel_norm_gs = np.linalg.norm(self.agent_gs.observed_trajectory[:, 2:4], axis=1)
        if self.sim_type == 'nds':
            vel_nds_vel_norm_gs = np.linalg.norm(self.gs_actual_trj[:, 2:4], axis=1)
            gs_nds_trj = np.array(self.gs_actual_trj[:, 0:2])
            gs_nds_heading = np.array(self.gs_actual_trj[:, 4]) / math.pi * 180

        num_frame = len(lt_ob_trj)

        "---- show plans at each time step ----"
        axes[0].plot(cv_1[:, 0], cv_1[:, 1], 'b-', linewidth=0.1)
        axes[0].plot(cv_2[:, 0], cv_2[:, 1], 'r-', linewidth=0.1)

        # ----position at each time step
        # version 1
        # for t in range(num_frame):
        #     # simulation results
        #     draw_rectangle(lt_ob_trj[t, 0], lt_ob_trj[t, 1], lt_ob_heading[t], axes[0],
        #                    para_alpha=0.3, para_color='#0E76CF')
        #     draw_rectangle(gs_ob_trj[t, 0], gs_ob_trj[t, 1], gs_ob_heading[t], axes[0],
        #                    para_alpha=0.3, para_color='#7030A0')
        #     #
        #     # nds ground truth
        #     if self.sim_type == 'nds':
        #         draw_rectangle(lt_nds_trj[t, 0], lt_nds_trj[t, 1], lt_nds_heading[t], axes[0],
        #                        para_alpha=0.3, para_color='blue')
        #
        #         draw_rectangle(gs_nds_trj[t, 0], gs_nds_trj[t, 1], gs_nds_heading[t], axes[0],
        #                        para_alpha=0.3, para_color='red')

        # version 2
        axes[0].scatter(lt_ob_trj[:num_frame, 0],
                        lt_ob_trj[:num_frame, 1],
                        s=50,
                        alpha=0.6,
                        color='#0E76CF',
                        label='left-turn simulation')
        axes[0].scatter(gs_ob_trj[:num_frame, 0],
                        gs_ob_trj[:num_frame, 1],
                        s=50,
                        alpha=0.6,
                        color='#7030A0',
                        label='go-straight simulation')

        if self.sim_type == 'nds':
            axes[0].scatter(lt_nds_trj[:num_frame, 0],
                            lt_nds_trj[:num_frame, 1],
                            s=50,
                            alpha=0.3,
                            color='blue',
                            label='left-turn NDS')
            axes[0].scatter(gs_nds_trj[:num_frame, 0],
                            gs_nds_trj[:num_frame, 1],
                            s=50,
                            alpha=0.3,
                            color='red',
                            label='go-straight NDS')

        # ----full tracks at each time step
        # for t in range(self.num_step):
        #     if int(t / 2) == t / 2:
        #         lt_track = self.agent_lt.trj_solution_collection[t]
        #         axes[0].plot(lt_track[:, 0], lt_track[:, 1], '--', color='black', alpha=0.5)
        #         gs_track = self.agent_gs.trj_solution_collection[t]
        #         axes[0].plot(gs_track[:, 0], gs_track[:, 1], '--', color='black', alpha=0.5)
        # if self.agent_gs.conl_type in {'gt', 'linear-gt'}:
        #     gs_inter_track = self.agent_gs.estimated_inter_agent.trj_solution_collection[t]
        #     axes[0].plot(gs_inter_track[:, 0], gs_inter_track[:, 1], '--', color='red', alpha=0.5)
        # if self.agent_lt.conl_type in {'gt', 'linear-gt'}:
        #     lt_inter_track = self.agent_lt.estimated_inter_agent.trj_solution_collection[t]
        #     axes[0].plot(lt_inter_track[:, 0], lt_inter_track[:, 1], '--', color='red', alpha=0.5)

        # ----connect two agents
        for t in range(self.num_step + 1):
            axes[0].plot([self.agent_lt.observed_trajectory[t, 0], self.agent_gs.observed_trajectory[t, 0]],
                         [self.agent_lt.observed_trajectory[t, 1], self.agent_gs.observed_trajectory[t, 1]],
                         color='gray',
                         alpha=0.2)

        "---- show velocity ----"
        x_range = np.array(range(np.size(self.agent_lt.observed_trajectory, 0)))
        axes[1].plot(x_range, vel_ob_vel_norm_lt, linestyle='--',
                     color='blue', label='left-turn simulation')
        axes[1].plot(x_range, vel_ob_vel_norm_gs, linestyle='--',
                     color='red', label='go-straight simulation')

        if self.sim_type == 'nds':
            x_range_nds = self.simu_time + x_range
            axes[1].plot(x_range, vel_nds_vel_norm_gs[x_range_nds],
                         color='red', label='go-straight NDS')
            axes[1].plot(x_range, vel_nds_vel_norm_lt[x_range_nds],
                         color='blue', label='left-turn NDS')

        axes[1].legend()
        axes[0].legend()
        # axes[0].axis('equal')
        plt.show()
        plt.savefig(file_path + self.tag + '-final.png', dpi=600)
        # plt.savefig(file_path + self.tag + '-final.svg')
        plt.close()

    def visualize_single_step(self, file_path):

        cv_lt = []
        cv_gs = []
        # set figures
        fig, axes = plt.subplots(1, 1, figsize=[5, 5])
        # fig.suptitle('trajectory_LT_' + self.semantic_result)
        axes.set_title('trajectory')
        if self.sim_type == 'simu_left_turn':
            axes.set(aspect=1, xlim=(-9.1, 24.9), ylim=(-13, 8))
            img = plt.imread('background_pic/T_intersection.jpg')
            axes.imshow(img, extent=[-9.1, 24.9, -13, 8])
            # central vertices
            cv_lt, _ = get_central_vertices('lt')
            cv_gs, _ = get_central_vertices('gs')
        elif self.sim_type == 'simu_ramp':
            cv_it, _ = get_central_vertices('ml')
            cv_gs, _ = get_central_vertices('ir')
        elif self.sim_type == 'nds':
            axes.set(aspect=1, xlim=(-22 - 13, 53 - 13), ylim=(-31 - 7.8, 57 - 7.8))
            img = plt.imread('background_pic/Jianhexianxia-v2.png')
            axes.imshow(img, extent=[-28 - 13, 58 - 13, -42 - 7.8, 64 - 7.8])
            # central vertices
            lt_origin_point = self.agent_lt.observed_trajectory[0, 0:2]
            gs_origin_point = self.agent_gs.observed_trajectory[0, 0:2]
            cv_lt, _ = get_central_vertices('lt_nds', origin_point=lt_origin_point)
            cv_gs, _ = get_central_vertices('gs_nds', origin_point=gs_origin_point)

        "---- data abstraction ----"
        # lt track (observed in simulation and ground truth in nds)
        lt_ob_trj = self.agent_lt.observed_trajectory[:, 0:2]
        # gs track (observed in simulation and ground truth in nds)
        gs_ob_trj = self.agent_gs.observed_trajectory[:, 0:2]

        # solution at this single time point
        lt_plan = self.agent_lt.trj_solution_collection[0]
        lt_inter_plan = self.agent_lt.estimated_inter_agent.trj_solution_collection[0]
        gs_plan = self.agent_gs.trj_solution_collection[0]
        gs_inter_plan = self.agent_gs.estimated_inter_agent.trj_solution_collection[0]

        # find the closest point pairs
        dis_lt_plan = np.linalg.norm(lt_plan[:, 0:2] - lt_inter_plan[:, 0:2], axis=1)
        lt_index_min_dis = np.where(dis_lt_plan == min(dis_lt_plan))

        dis_gs_plan = np.linalg.norm(gs_plan[:, 0:2] - gs_inter_plan[:, 0:2], axis=1)
        gs_index_min_dis = np.where(dis_gs_plan == min(dis_gs_plan))

        "---- show initial state ----"
        axes.plot(cv_lt[:, 0], cv_lt[:, 1], 'b-', linewidth=0.1)
        axes.plot(cv_gs[:, 0], cv_gs[:, 1], 'r-', linewidth=0.1)

        axes.scatter(lt_ob_trj[0, 0],
                     lt_ob_trj[0, 1],
                     s=50,
                     alpha=0.6,
                     color='#0E76CF',
                     label='left-turn-ipv-' + str(self.agent_lt.ipv))
        axes.scatter(gs_ob_trj[0, 0],
                     gs_ob_trj[0, 1],
                     s=50,
                     alpha=0.6,
                     color='#7030A0',
                     label='go-straight-ipv-' + str(self.agent_gs.ipv))

        # ----full tracks at each time step
        axes.scatter(lt_plan[:, 0], lt_plan[:, 1], color='blue', alpha=0.5, s=3)
        axes.scatter(gs_plan[:, 0], gs_plan[:, 1], color='red', alpha=0.5, s=3)
        if self.agent_gs.conl_type in {'gt', 'linear-gt'}:
            axes.scatter(gs_inter_plan[:, 0], gs_inter_plan[:, 1], color='red', alpha=0.2, s=3)
        if self.agent_lt.conl_type in {'gt', 'linear-gt'}:
            axes.scatter(lt_inter_plan[:, 0], lt_inter_plan[:, 1], color='blue', alpha=0.2, s=3)

        # ----connect two agents
        lt_index = int(lt_index_min_dis[0])
        for t in {lt_index}:
            axes.plot([lt_plan[t, 0], lt_inter_plan[t, 0]],
                      [lt_plan[t, 1], lt_inter_plan[t, 1]],
                      color='gray',
                      alpha=0.4,
                      linewidth=0.2)
        gs_index = int(gs_index_min_dis[0])
        for t in {gs_index}:
            axes.plot([gs_plan[t, 0], gs_inter_plan[t, 0]],
                      [gs_plan[t, 1], gs_inter_plan[t, 1]],
                      color='purple',
                      alpha=0.4,
                      linewidth=0.2)

        # plt.show()
        axes.legend()
        plt.savefig(file_path + self.tag, dpi=300)
        # plt.savefig(file_path + self.tag + '-final.svg')
        # plt.close()

    def visualize_multi_interaction(self, trj_coll, t, file_path='./'):

        fig_num_sqrt = int(np.sqrt(len(trj_coll[0].values())))

        # central vertices
        lt_origin_point = self.agent_lt.observed_trajectory[0, 0:2]
        gs_origin_point = self.agent_gs.observed_trajectory[0, 0:2]
        cv_lt, _ = get_central_vertices('lt_nds', origin_point=lt_origin_point)
        cv_gs, _ = get_central_vertices('gs_nds', origin_point=gs_origin_point)

        # set figures
        fig, axes = plt.subplots(fig_num_sqrt + 1, fig_num_sqrt, figsize=[12, 12])

        # ground truth trajectory
        act_trj_lt = self.lt_actual_trj[:, 0:2]
        act_trj_gs = self.gs_actual_trj[:, 0:2]
        axes[0, 2].plot(act_trj_lt[:, 0], act_trj_lt[:, 1], 'b-', linewidth=2)
        axes[0, 2].plot(act_trj_gs[:, 0], act_trj_gs[:, 1], 'r-', linewidth=2)
        # axes[0, 2].set(aspect=1, xlim=(-7.2, 24), ylim=(0, 32))
        axes[0, 2].set(aspect=1,
                       xlim=(min(np.concatenate([act_trj_lt[:, 0], act_trj_gs[:, 0]])) - 4,
                             max(np.concatenate([act_trj_lt[:, 0], act_trj_gs[:, 0]])) + 4),
                       ylim=(min(np.concatenate([act_trj_lt[:, 1], act_trj_gs[:, 1]])) - 4,
                             max(np.concatenate([act_trj_lt[:, 1], act_trj_gs[:, 1]])) + 4))
        img = plt.imread('background_pic/Jianhexianxia-v2.png')
        axes[0, 2].imshow(img, extent=[-28 - 13, 58 - 13, -42 - 7.8, 64 - 7.8])

        simu_trj_coll = trj_coll[t]
        simu_opt_trj = simu_trj_coll['non-interactive']
        compare_len = min(np.size(act_trj_lt, 0), np.size(simu_opt_trj['lt'], 0))
        compare_range = range(compare_len)

        for task_id in range(len(simu_trj_coll) - 1):
            # lt track (observed in simulation and ground truth in nds)
            lt_plan = simu_trj_coll['task' + str(task_id)]['lt-self']
            lt_inter_plan = simu_trj_coll['task' + str(task_id)]['lt-inter']

            # gs track (observed in simulation and ground truth in nds)
            gs_plan = simu_trj_coll['task' + str(task_id)]['gs-self']
            gs_inter_plan = simu_trj_coll['task' + str(task_id)]['gs-inter']

            axe_col = int(task_id / fig_num_sqrt) + 1
            axe_row = int(task_id % fig_num_sqrt)
            axe = axes[axe_col, axe_row]

            #  set ipv of LT vehicle
            case_ipv = simu_trj_coll['task' + str(task_id)]['ipv']
            axe.set_title(str(task_id) + '-ipv-lt: ' + str(case_ipv[0]) + 'gs: ' + str(case_ipv[1]))

            axe.set(aspect=1,
                    xlim=(min(np.concatenate([act_trj_lt[:, 0], act_trj_gs[:, 0]])) - 4,
                          max(np.concatenate([act_trj_lt[:, 0], act_trj_gs[:, 0]])) + 4),
                    ylim=(min(np.concatenate([act_trj_lt[:, 1], act_trj_gs[:, 1]])) - 4,
                          max(np.concatenate([act_trj_lt[:, 1], act_trj_gs[:, 1]])) + 4))
            img = plt.imread('background_pic/Jianhexianxia-v2.png')
            axe.imshow(img, extent=[-28 - 13, 58 - 13, -42 - 7.8, 64 - 7.8])

            "---- reference ----"
            axe.plot(cv_lt[:, 0], cv_lt[:, 1], 'b-', linewidth=0.1)
            axe.plot(cv_gs[:, 0], cv_gs[:, 1], 'r-', linewidth=0.1)

            # ----full tracks at each time step
            axe.scatter(lt_plan[compare_range, 0], lt_plan[compare_range, 1], color='blue', alpha=0.5, s=3)
            axe.scatter(gs_plan[compare_range, 0], gs_plan[compare_range, 1], color='red', alpha=0.5, s=3)
            axe.plot(act_trj_lt[compare_range, 0], act_trj_lt[compare_range, 1], 'gray', linewidth=2)
            axe.plot(act_trj_gs[compare_range, 0], act_trj_gs[compare_range, 1], 'gray', linewidth=2)

            # if self.agent_gs.conl_type in {'gt', 'linear-gt'}:
            #     axe.scatter(gs_inter_plan[:, 0], gs_inter_plan[:, 1], color='gray', alpha=0.5, s=3)
            # if self.agent_lt.conl_type in {'gt', 'linear-gt'}:
            #     axe.scatter(lt_inter_plan[:, 0], lt_inter_plan[:, 1], color='gray', alpha=0.5, s=3)
        plt.tight_layout()
        plt.show()
        # # solution at this single time point
        # lt_plan = self.agent_lt.trj_solution_collection[0]
        # lt_inter_plan = self.agent_lt.estimated_inter_agent[0].trj_solution_collection[0]
        # gs_plan = self.agent_gs.trj_solution_collection[0]
        # gs_inter_plan = self.agent_gs.estimated_inter_agent[0].trj_solution_collection[0]
        #
        # # find the closest point pairs
        # dis_lt_plan = np.linalg.norm(lt_plan[:, 0:2] - lt_inter_plan[:, 0:2], axis=1)
        # lt_index_min_dis = np.where(dis_lt_plan == min(dis_lt_plan))
        #
        # dis_gs_plan = np.linalg.norm(gs_plan[:, 0:2] - gs_inter_plan[:, 0:2], axis=1)
        # gs_index_min_dis = np.where(dis_gs_plan == min(dis_gs_plan))
        #
        # "---- show initial state ----"
        # axes.plot(cv_lt[:, 0], cv_lt[:, 1], 'b-', linewidth=0.1)
        # axes.plot(cv_gs[:, 0], cv_gs[:, 1], 'r-', linewidth=0.1)
        #
        # axes.scatter(act_trj_lt[0, 0],
        #              act_trj_lt[0, 1],
        #              s=50,
        #              alpha=0.6,
        #              color='#0E76CF',
        #              label='left-turn-ipv-' + str(self.agent_lt.ipv))
        # axes.scatter(gs_ob_trj[0, 0],
        #              gs_ob_trj[0, 1],
        #              s=50,
        #              alpha=0.6,
        #              color='#7030A0',
        #              label='go-straight-ipv-' + str(self.agent_gs.ipv))
        #
        # # ----connect two agents
        # lt_index = int(lt_index_min_dis[0])
        # for t in {lt_index}:
        #     axes.plot([lt_plan[t, 0], lt_inter_plan[t, 0]],
        #               [lt_plan[t, 1], lt_inter_plan[t, 1]],
        #               color='gray',
        #               alpha=0.4,
        #               linewidth=0.2)
        # gs_index = int(gs_index_min_dis[0])
        # for t in {gs_index}:
        #     axes.plot([gs_plan[t, 0], gs_inter_plan[t, 0]],
        #               [gs_plan[t, 1], gs_inter_plan[t, 1]],
        #               color='purple',
        #               alpha=0.4,
        #               linewidth=0.2)
        #
        # # plt.show()
        # axes.legend()
        # plt.savefig(file_path + self.tag, dpi=300)
        # # plt.savefig(file_path + self.tag + '-final.svg')
        # plt.close()

    def read_nds_scenario(self, controller_type_lt, controller_type_gs, t=0):
        cross_id, data_cross, _ = analyze_ipv_in_nds(self.case_id)
        # data_cross:
        # 0-ipv_lt | ipv_lt_error | lt_px | lt_py  | lt_vx  | lt_vy  | lt_heading  |...
        # 7-ipv_gs | ipv_gs_error | gs_px | gs_py  | gs_vx  | gs_vy  | gs_heading  |

        if cross_id == -1:
            print('-----no trajectory crossed in case' + str(self.case_id))
            return None
        else:
            init_position_lt = [data_cross[t, 2] - 13, data_cross[t, 3] - 7.8]
            init_velocity_lt = [data_cross[t, 4], data_cross[t, 5]]
            init_acceleration_lt = [(data_cross[t + 1, 4] - data_cross[t + 1, 4]) / 0.12,
                                    (data_cross[t + 1, 5] - data_cross[t + 1, 5]) / 0.12]
            init_heading_lt = data_cross[t, 6]
            if controller_type_lt in {'opt', 'gt', 'linear-gt'}:
                ipv_weight_lt = 1 - data_cross[4:, 1]
                ipv_weight_lt = ipv_weight_lt / ipv_weight_lt.sum()
                ipv_lt = sum(ipv_weight_lt * data_cross[4:, 0])
                # ipv_lt = max(sum(ipv_weight_lt * data_cross[4:, 0])-0.2, -math.pi*3/8)
            else:
                ipv_lt = 0

            init_position_gs = [data_cross[t, 9] - 13, data_cross[t, 10] - 7.8]
            init_velocity_gs = [data_cross[t, 11], data_cross[t, 12]]
            init_acceleration_gs = [(data_cross[t + 1, 11] - data_cross[t + 1, 11]) / 0.12,
                                    (data_cross[t + 1, 12] - data_cross[t + 1, 12]) / 0.12]
            init_heading_gs = data_cross[t, 13]
            if controller_type_gs in {'opt', 'gt', 'linear-gt'}:
                ipv_weight_gs = 1 - data_cross[4:, 8]
                ipv_weight_gs = ipv_weight_gs / ipv_weight_gs.sum()
                ipv_gs = sum(ipv_weight_gs * data_cross[4:, 7])
                # ipv_gs = max(sum(ipv_weight_gs * data_cross[4:, 7])-0.2, -math.pi*3/8)
            else:
                ipv_gs = 0
            self.lt_actual_trj = data_cross[:, 2:7]
            self.lt_actual_trj[:, 0] = self.lt_actual_trj[:, 0] - 13
            self.lt_actual_trj[:, 1] = self.lt_actual_trj[:, 1] - 7.8

            self.gs_actual_trj = data_cross[:, 9:14]
            self.gs_actual_trj[:, 0] = self.gs_actual_trj[:, 0] - 13
            self.gs_actual_trj[:, 1] = self.gs_actual_trj[:, 1] - 7.8

            self.case_len = np.size(data_cross, 0) - 1

            return Scenario([init_position_lt, init_position_gs],
                            [init_velocity_lt, init_velocity_gs],
                            [init_heading_lt, init_heading_gs],
                            [ipv_lt, ipv_gs],
                            [controller_type_lt, controller_type_gs])

    def read_nds_scenario_multi(self, controller_type_lt, controller_type_gs, t=0):
        mat = scipy.io.loadmat('./data/NDS_data_fixed.mat')
        # full interaction information
        inter_info = mat['interaction_info']
        inter_num = mat['interact_agent_num']
        case_info = inter_info[self.case_id]

        # find co-present gs agents (not necessarily interacting)
        gs_info_multi = case_info[1:inter_num[0, self.case_id] + 1]
        gs_num = len(gs_info_multi)

        lt_info = case_info[0]
        gs_info = gs_info_multi[self.gs_id]
        solid_frame = np.nonzero(gs_info[:, 0])[0]
        solid_range = range(solid_frame[0], solid_frame[-1])
        t += solid_frame[0]

        init_position_lt = [lt_info[t, 0] - 13, lt_info[t, 1] - 7.8]
        init_velocity_lt = [lt_info[t, 2], lt_info[t, 3]]
        init_acceleration_lt = [(lt_info[t + 1, 2] - lt_info[t + 1, 3]) / 0.12,
                                (lt_info[t + 1, 2] - lt_info[t + 1, 3]) / 0.12]
        init_heading_lt = lt_info[t, 4]

        init_position_gs = [gs_info[t, 0] - 13, gs_info[t, 1] - 7.8]
        init_velocity_gs = [gs_info[t, 2], gs_info[t, 3]]
        init_acceleration_gs = [(gs_info[t + 1, 2] - gs_info[t + 1, 3]) / 0.12,
                                (gs_info[t + 1, 2] - gs_info[t + 1, 3]) / 0.12]
        init_heading_gs = gs_info[t, 4]

        self.lt_actual_trj = lt_info[solid_range, 0:2]
        self.lt_actual_trj[:, 0] = self.lt_actual_trj[:, 0] - 13
        self.lt_actual_trj[:, 1] = self.lt_actual_trj[:, 1] - 7.8

        self.gs_actual_trj = gs_info[solid_range, 0:2]
        self.gs_actual_trj[:, 0] = self.gs_actual_trj[:, 0] - 13
        self.gs_actual_trj[:, 1] = self.gs_actual_trj[:, 1] - 7.8

        self.case_len = len(solid_range)

        return Scenario([init_position_lt, init_position_gs],
                        [init_velocity_lt, init_velocity_gs],
                        [init_heading_lt, init_heading_gs],
                        [0, 0],
                        [controller_type_lt, controller_type_gs])

    def ipv_analysis(self):
        """
        estimate self ipv expression from a third-party view
        (On answering what is my ipv like from the view of others?)

        note that due to the dependence on [last self track], this process is not possible when using linearized
        optimization.
        Returns -------

        """
        ipv_collection_lt = []
        ipv_collection_gs = []
        ipv_error_collection_lt = []
        ipv_error_collection_gs = []
        track_lt = self.agent_lt.observed_trajectory
        track_gs = self.agent_gs.observed_trajectory
        for t in range(np.size(track_lt, 0) - 6):
            init_position_lt = track_lt[t, 0:2]
            init_velocity_lt = track_lt[t, 2:4]
            init_heading_lt = track_lt[t, 4]
            agent_lt = Agent(init_position_lt, init_velocity_lt, init_heading_lt, 'lt_nds')
            track_lt_temp = track_lt[t:t + 6, 0:2]

            init_position_gs = track_gs[t, 0:2]
            init_velocity_gs = track_gs[t, 2:4]
            init_heading_gs = track_gs[t, 4]
            agent_gs = Agent(init_position_gs, init_velocity_gs, init_heading_gs, 'gs_nds')
            track_gs_temp = track_gs[t:t + 6, 0:2]

            # estimate ipv
            agent_lt.estimate_self_ipv(track_lt_temp, track_gs_temp)
            ipv_collection_lt.append(agent_lt.ipv)
            ipv_error_collection_lt.append(agent_lt.ipv_error)

            agent_gs.estimate_self_ipv(track_gs_temp, track_lt_temp)
            ipv_collection_gs.append(agent_gs.ipv)
            ipv_error_collection_gs.append(agent_gs.ipv_error)

        # figure preparation
        x_range = np.array(range(len(ipv_collection_lt)))
        ipv_collection_lt = np.array(ipv_collection_lt)
        ipv_collection_gs = np.array(ipv_collection_gs)
        ipv_error_collection_lt = np.array(ipv_error_collection_lt)
        ipv_error_collection_gs = np.array(ipv_error_collection_gs)

        plt.figure()
        plt.plot(x_range, ipv_collection_lt, color='#0E76CF')
        plt.fill_between(x_range,
                         ipv_collection_lt - ipv_error_collection_lt,
                         ipv_collection_lt + ipv_error_collection_lt,
                         alpha=0.3,
                         color='#0E76CF',
                         label='estimated lt IPV')
        plt.plot(x_range, ipv_collection_gs, color='#7030A0')
        plt.fill_between(x_range,
                         ipv_collection_gs - ipv_error_collection_gs,
                         ipv_collection_gs + ipv_error_collection_gs,
                         alpha=0.3,
                         color='#7030A0',
                         label='estimated gs IPV')

        # y_lt = np.array(ipv_collection_lt)
        # point_smoothed_lt, _ = smooth_ployline(np.array([x_range, y_lt]).T)
        # x_smoothed_lt = point_smoothed_lt[:, 0]
        # y_lt_smoothed = point_smoothed_lt[:, 1]
        # y_error_lt = np.array(ipv_error_collection_lt)
        # error_smoothed_lt, _ = smooth_ployline(np.array([x_range, y_error_lt]).T)
        # y_error_lt_smoothed = error_smoothed_lt[:, 1]
        #
        # y_gs = np.array(ipv_collection_gs)
        # point_smoothed_gs, _ = smooth_ployline(np.array([x_range, y_gs]).T)
        # x_smoothed_gs = point_smoothed_gs[:, 0]
        # y_gs_smoothed = point_smoothed_gs[:, 1]
        # y_error_gs = np.array(ipv_error_collection_gs)
        # error_smoothed_gs, _ = smooth_ployline(np.array([x_range, y_error_gs]).T)
        # y_error_gs_smoothed = error_smoothed_gs[:, 1]
        #
        # plt.figure()
        # plt.fill_between(x_smoothed_lt, y_lt_smoothed - y_error_lt_smoothed, y_lt_smoothed + y_error_lt_smoothed,
        #                  alpha=0.3,
        #                  color='blue',
        #                  label='estimated lt IPV')
        # plt.fill_between(x_smoothed_gs, y_gs_smoothed - y_error_gs_smoothed, y_gs_smoothed + y_error_gs_smoothed,
        #                  alpha=0.3,
        #                  color='red',
        #                  label='estimated gs IPV')

    def cal_trj_similarity(self, trj_coll, isFig=False):

        # Ground truth trajectory
        act_trj_lt = self.lt_actual_trj[:, 0:2]
        act_trj_gs = self.gs_actual_trj[:, 0:2]

        # Simulated trajectory
        simu_trj_coll = trj_coll[0]
        simu_opt_trj = simu_trj_coll['non-interactive']
        case_num_sqrt = int(np.sqrt(len(simu_trj_coll.values())))

        compare_len = min(np.size(act_trj_lt, 0), np.size(simu_opt_trj['lt'], 0))
        compare_range = range(compare_len)

        " Calculate similarity between simulation and ground truth with MLE "
        act_trj = np.concatenate((act_trj_lt[compare_range, :], act_trj_gs[compare_range, :]), axis=0)
        similarity = np.zeros(len(simu_trj_coll) - 1)
        for i in range(len(simu_trj_coll) - 1):
            simu_track = np.concatenate((simu_trj_coll['task' + str(i)]['lt-self'][compare_range, :],
                                         simu_trj_coll['task' + str(i)]['gs-self'][compare_range, :]), axis=0)
            rel_dis = np.linalg.norm(simu_track - act_trj, axis=1)  # distance vector
            # print('case', str(i))
            # print('max dis:', max(rel_dis))
            # print('ave dis:', np.mean(rel_dis))
            # print('----')
            sim_factor = np.exp(- rel_dis ** 2 / (2 * sigma ** 2))
            similarity[i] = np.prod((1 / sigma / np.sqrt(2 * math.pi)) * sim_factor) ** (1 / np.size(act_trj, 0))

            # Set similarity to 0 if it is negative
            similarity[i] = max(similarity[i], 0)
        similarity = similarity / sum(similarity)
        mean_ipv_lt = 0
        mean_ipv_gs = 0
        for i in range(len(simu_trj_coll) - 1):
            mean_ipv_lt += simu_trj_coll['task' + str(i)]['ipv'][0] * similarity[i]
            mean_ipv_gs += simu_trj_coll['task' + str(i)]['ipv'][1] * similarity[i]
        similarity = similarity.reshape([case_num_sqrt, case_num_sqrt])

        # visualize similarity matrix
        if isFig:
            _, axes = plt.subplots(1, 2, figsize=[20, 10])
            axes[0].set_title('NDS similarity')
            sns.heatmap(similarity, annot=True, cmap='Blues', ax=axes[0])
            # plt.show()
            return axes, [mean_ipv_lt, mean_ipv_gs]
        return None, [mean_ipv_lt, mean_ipv_gs]

    def cal_interaction_strength(self, trj_coll, ax=None, isFig=True, isSaveFig=False, file_dir='./'):

        # Ground truth trajectory
        act_trj_lt = self.lt_actual_trj[:, 0:2]
        act_trj_gs = self.gs_actual_trj[:, 0:2]
        interaction_distance = min(np.linalg.norm(act_trj_lt - act_trj_gs, axis=1)) - 1

        # Simulated trajectory
        simu_trj_coll = trj_coll[0]
        simu_opt_trj = simu_trj_coll['non-interactive']
        case_num_sqrt = int(np.sqrt(len(simu_trj_coll.values())))

        compare_len = min(np.size(act_trj_lt, 0), np.size(simu_opt_trj['lt'], 0))
        compare_range = range(compare_len)

        dyna_mat = [2 * i + 1 for i in compare_range]
        dyna_mat = np.array([dyna_mat, dyna_mat])
        dyna_mat = dyna_mat.T

        opt_trj = np.concatenate((simu_opt_trj['lt'][compare_range, :], simu_opt_trj['gs'][compare_range, :]), axis=0)
        inter_strength = np.zeros(len(simu_trj_coll) - 1)

        " Quantify interaction strength by compare simulated trajectory and optimal one "
        # for i in range(len(simu_trj_coll) - 1):
        #     simu_track = np.concatenate((simu_trj_coll['task' + str(i)]['lt-self'][compare_range, :],
        #                                  simu_trj_coll['task' + str(i)]['gs-self'][compare_range, :]), axis=0)
        #     rel_dis = np.linalg.norm(simu_track - opt_trj, axis=1)  # distance vector
        #     # print('case', str(i))
        #     # print('max dis:', max(rel_dis))
        #     # print('ave dis:', np.mean(rel_dis))
        #     # print('----')
        #     sim_factor = np.exp(- rel_dis ** 2 / (2 * sigma ** 2))
        #     inter_strength[i] = np.prod((1 / sigma / np.sqrt(2 * math.pi)) * sim_factor) ** (1 / np.size(opt_trj, 0))
        #
        #     # Set inter_strength to 0 if it is negative
        #     inter_strength[i] = max(inter_strength[i], 0)

        " Quantify interaction strength by sensitivity analysis "
        for i in range(len(simu_trj_coll) - 1):
            lt_plan = simu_trj_coll['task' + str(i)]['lt-self'][compare_range, :]
            gs_plan = simu_trj_coll['task' + str(i)]['gs-self'][compare_range, :]
            plan_vec_lt = np.array(lt_plan - gs_plan)
            plan_dis_lt = np.linalg.norm(plan_vec_lt, axis=1)
            plan_vec_lt[np.where(plan_dis_lt > interaction_distance), :] = 0
            sensi_r2p = plan_vec_lt / np.linalg.norm(plan_dis_lt) * 0.5 * 0.12 ** 2
            inter_strength[i] = np.sum(abs(dyna_mat * sensi_r2p))

        inter_strength = inter_strength.reshape([case_num_sqrt, case_num_sqrt])

        # visualize inter_strength matrix
        if isFig:
            ax[1].set_title('interaction strength')
            sns.heatmap(inter_strength, annot=True, cmap='Reds', ax=ax[1])
            if isSaveFig:
                plt.savefig(file_dir + 'figure/' + str(self.case_id), dpi=450)
                plt.close()
            # plt.show()
        return inter_strength

    def min_inter_strength_test(self, estimated_ipv, inter_str_matrix, isSave=False, file_name=None):
        """
        check whether the min-interaction-strength case is the most human-like one
        Returns
        -------

        """

        # find the most likely case according to mean ipv
        lt_id = None
        gs_id = None
        for i in range(len(self.ipv_list) - 1):
            if self.ipv_list[i] < estimated_ipv[0] < self.ipv_list[i + 1]:
                lt_id = i
            if self.ipv_list[i] < estimated_ipv[1] < self.ipv_list[i + 1]:
                gs_id = i

        actual_inter_str = inter_str_matrix[lt_id, gs_id]
        min_inter_str = inter_str_matrix.min()
        mean_inter_str = np.mean(inter_str_matrix)
        std_inter_str = np.std(inter_str_matrix)
        print('actual interaction strength: ', actual_inter_str)
        print('minimal interaction strength: ', min_inter_str)
        print('mean interaction strength: ', mean_inter_str)
        print('std interaction strength: ', std_inter_str)

        if isSave:
            "---- sava data ----"
            # prepare data
            df = pd.DataFrame([[self.case_id,
                                actual_inter_str, min_inter_str,
                                mean_inter_str, std_inter_str,
                                ], ],
                              columns=['case id',
                                       'actual interaction strength', 'minimal interaction strength',
                                       'mean interaction strength', 'std interaction strength',
                                       ])

            # write data
            if self.case_id == 0:
                header_flag = True
                start_row = 0
            else:
                header_flag = False
                start_row = self.case_id + 1

            with pd.ExcelWriter(file_name, mode='a', if_sheet_exists="overlay", engine="openpyxl") as writer:
                df.to_excel(writer, header=header_flag, index=False,
                            startcol=0, startrow=start_row)

    def save_test_meta(self, param_v, ipv, file_name, sheet_name):
        """

        Returns
        -------

        """

        "---- event data abstraction ----"
        # lt track (observed in simulation and ground truth in nds)
        ob_trj_1 = self.agent_lt.observed_trajectory[:, 0:2]
        vel_ob_vel_norm_1 = np.linalg.norm(self.agent_lt.observed_trajectory[:, 2:4], axis=1)
        acc_ob_1 = (vel_ob_vel_norm_1[1:] - vel_ob_vel_norm_1[:-1]) / 0.12
        jerk_ob_1 = (acc_ob_1[1:] - acc_ob_1[:-1]) / 0.12

        # gs track (observed in simulation and ground truth in nds)
        ob_trj_2 = self.agent_gs.observed_trajectory[:, 0:2]
        vel_ob_vel_norm_2 = np.linalg.norm(self.agent_gs.observed_trajectory[:, 2:4], axis=1)
        acc_ob_2 = (vel_ob_vel_norm_2[1:] - vel_ob_vel_norm_2[:-1]) / 0.12
        jerk_ob_2 = (acc_ob_2[1:] - acc_ob_2[:-1]) / 0.12

        "---- meta data ----"
        # 1
        case_id = self.case_id

        # ----semantic result
        seman_res_simu = self.semantic_result

        # ----APET
        apet_simu, _, _ = cal_pet(ob_trj_1, ob_trj_2, type_cal='apet')
        # 15
        min_apet_simu = apet_simu.min()
        # 16
        mean_apet_simu = min(apet_simu.mean(), 15)

        # ----PET
        # 18
        pet_simu, _ = cal_pet(ob_trj_1, ob_trj_2, type_cal='pet')

        # ----max acc and jerk
        # 19
        max_acc_simu_lt = max(max(acc_ob_1), -min(acc_ob_1))
        # 20
        max_acc_simu_gs = max(max(acc_ob_2), -min(acc_ob_2))

        # 23
        max_jerk_simu_lt = max(max(jerk_ob_1), -min(jerk_ob_1))
        # 24
        max_jerk_simu_gs = max(max(jerk_ob_2), -min(jerk_ob_2))

        "---- sava data ----"
        # prepare data
        df = pd.DataFrame([[case_id,
                            param_v, ipv,
                            seman_res_simu,
                            min_apet_simu, mean_apet_simu, pet_simu,
                            max_acc_simu_lt, max_acc_simu_gs,
                            max_jerk_simu_lt, max_jerk_simu_gs,
                            ], ],
                          columns=['case id',
                                   'param_v', 'ipv',
                                   'seman res',
                                   'min apet', 'mean apet', 'pet',
                                   'MAX acc. lt', 'MAX acc. gs',
                                   'MAX jerk lt', 'MAX jerk gs',
                                   ])

        # write data
        if case_id == 0:
            header_flag = True
            start_row = 0
        else:
            header_flag = False
            start_row = case_id + 1

        with pd.ExcelWriter(file_name, mode='a', if_sheet_exists="overlay", engine="openpyxl") as writer:
            df.to_excel(writer, header=header_flag, index=False,
                        sheet_name=sheet_name,
                        startcol=0, startrow=start_row)

    def save_simu_meta(self, num_failed, file_name, sheet_name):
        """

        Returns
        -------

        """

        "---- event data abstraction ----"
        # lt track (observed in simulation and ground truth in nds)
        ob_trj_lt = self.agent_lt.observed_trajectory[:, 0:2]
        nds_trj_lt = np.array(self.lt_actual_trj[:, 0:2])
        vel_ob_vel_norm_lt = np.linalg.norm(self.agent_lt.observed_trajectory[:, 2:4], axis=1)
        vel_nds_vel_norm_lt = np.linalg.norm(self.lt_actual_trj[:, 2:4], axis=1)
        acc_ob_lt = (vel_ob_vel_norm_lt[1:] - vel_ob_vel_norm_lt[:-1]) / 0.12
        acc_nds_lt = (vel_nds_vel_norm_lt[1:] - vel_nds_vel_norm_lt[:-1]) / 0.12
        jerk_ob_lt = (acc_ob_lt[1:] - acc_ob_lt[:-1]) / 0.12
        jerk_nds_lt = (acc_nds_lt[1:] - acc_nds_lt[:-1]) / 0.12

        # gs track (observed in simulation and ground truth in nds)
        ob_trj_gs = self.agent_gs.observed_trajectory[:, 0:2]
        nds_trj_gs = np.array(self.gs_actual_trj[:, 0:2])
        vel_ob_vel_norm_gs = np.linalg.norm(self.agent_gs.observed_trajectory[:, 2:4], axis=1)
        vel_nds_vel_norm_gs = np.linalg.norm(self.gs_actual_trj[:, 2:4], axis=1)
        acc_ob_gs = (vel_ob_vel_norm_gs[1:] - vel_ob_vel_norm_gs[:-1]) / 0.12
        acc_nds_gs = (vel_nds_vel_norm_gs[1:] - vel_nds_vel_norm_gs[:-1]) / 0.12
        jerk_ob_gs = (acc_ob_gs[1:] - acc_ob_gs[:-1]) / 0.12
        jerk_nds_gs = (acc_nds_gs[1:] - acc_nds_gs[:-1]) / 0.12

        "---- meta data ----"
        # 1
        case_id = self.case_id

        # ----semantic result
        seman_res_simu = self.semantic_result
        seman_res_nds = get_semantic_result(nds_trj_lt, nds_trj_gs, case_type='nds')
        # 2
        seman_right = bool(seman_res_nds == seman_res_simu)

        # ----velocity
        # 3
        mean_vel_simu_lt = vel_ob_vel_norm_lt.mean()
        # 4
        mean_vel_nds_lt = vel_nds_vel_norm_lt.mean()
        # 5
        mean_vel_simu_gs = vel_ob_vel_norm_gs.mean()
        # 6
        mean_vel_nds_gs = vel_nds_vel_norm_gs.mean()
        # 7
        vel_rmse_lt = np.sqrt(
            ((vel_ob_vel_norm_lt - vel_nds_vel_norm_lt) ** 2).sum()
            / np.size(vel_ob_vel_norm_lt, 0))
        # 8
        vel_rmse_gs = np.sqrt(
            ((vel_ob_vel_norm_gs - vel_nds_vel_norm_gs) ** 2).sum()
            / np.size(vel_ob_vel_norm_gs, 0))

        # ----position deviation
        # 9
        ave_pos_dev_lt = np.linalg.norm(ob_trj_lt - nds_trj_lt, axis=1).mean()
        # 10
        pos_rmse_lt = np.sqrt(
            ((ave_pos_dev_lt - np.linalg.norm(ob_trj_lt - nds_trj_lt, axis=1)) ** 2).sum()
            / np.size(ob_trj_lt, 0))
        # 11
        ave_pos_dev_gs = np.linalg.norm(ob_trj_gs - nds_trj_gs, axis=1).mean()
        # 12
        pos_rmse_gs = np.sqrt(
            ((ave_pos_dev_gs - np.linalg.norm(ob_trj_gs - nds_trj_gs, axis=1)) ** 2).sum()
            / np.size(ob_trj_gs, 0))

        # ----APET
        apet_nds, _, _ = cal_pet(nds_trj_lt, nds_trj_gs, type_cal='apet')
        # 13
        min_apet_nds = apet_nds.min()
        # 14
        mean_apet_nds = min(apet_nds.mean(), 15)

        apet_simu, _, _ = cal_pet(ob_trj_lt, ob_trj_gs, type_cal='apet')
        # 15
        min_apet_simu = apet_simu.min()
        # 16
        mean_apet_simu = min(apet_simu.mean(), 15)

        # ----PET
        # 17
        pet_nds, _ = cal_pet(nds_trj_lt, nds_trj_gs, type_cal='pet')
        # 18
        pet_simu, _ = cal_pet(ob_trj_lt, ob_trj_gs, type_cal='pet')

        # ----max acc and jerk
        # 19
        max_acc_simu_lt = max(max(acc_ob_lt), -min(acc_ob_lt))
        # 20
        max_acc_simu_gs = max(max(acc_ob_gs), -min(acc_ob_gs))
        # 21
        max_acc_nds_lt = max(max(acc_nds_lt), -min(acc_nds_lt))
        # 22
        max_acc_nds_gs = max(max(acc_nds_gs), -min(acc_nds_gs))

        # 23
        max_jerk_simu_lt = max(max(jerk_ob_lt), -min(jerk_ob_lt))
        # 24
        max_jerk_simu_gs = max(max(jerk_ob_gs), -min(jerk_ob_gs))
        # 25
        max_jerk_nds_lt = max(max(jerk_nds_lt), -min(jerk_nds_lt))
        # 26
        max_jerk_nds_gs = max(max(jerk_nds_gs), -min(jerk_nds_gs))

        "---- sava data ----"
        # prepare data
        df = pd.DataFrame([[case_id, seman_right, seman_res_simu,
                            mean_vel_simu_lt, mean_vel_nds_lt,
                            mean_vel_simu_gs, mean_vel_nds_gs,
                            vel_rmse_lt, vel_rmse_gs,
                            ave_pos_dev_lt, pos_rmse_lt,
                            ave_pos_dev_gs, pos_rmse_gs,
                            min_apet_nds, min_apet_simu,
                            mean_apet_nds, mean_apet_simu,
                            pet_nds, pet_simu,
                            max_acc_simu_lt, max_acc_simu_gs,
                            max_acc_nds_lt, max_acc_nds_gs,
                            max_jerk_simu_lt, max_jerk_simu_gs,
                            max_jerk_nds_lt, max_jerk_nds_gs
                            ], ],
                          columns=['case id', 'semantic', ' result',
                                   'simu. v. lt', 'nds v. lt',
                                   'simu. v. gs', 'nds v. gs',
                                   'v. RMSE lt', 'velocity RMSE gs',
                                   'ave. POS deviation lt', 'POS RMSE lt',
                                   'ave. POS deviation gs', 'POS RMSE gs',
                                   'MIN APET NDS', 'MIN APET SIMU',
                                   'mean APET NDS', 'mean APET SIMU',
                                   'PET NDS', 'PET SIMU',
                                   'MAX acc. SIMU lt', 'MAX acc. SIMU gs',
                                   'MAX acc. NDS lt', 'MAX acc. NDS gs',
                                   'MAX jerk SIMU lt', 'MAX jerk SIMU gs',
                                   'MAX jerk NDS lt', 'MAX jerk NDS gs',
                                   ])

        # write data
        if case_id == 0:
            header_flag = True
            start_row = 0
        else:
            header_flag = False
            start_row = case_id + 1

        with pd.ExcelWriter(file_name, mode='a', if_sheet_exists="overlay", engine="openpyxl") as writer:
            df.to_excel(writer, header=header_flag, index=False,
                        sheet_name=sheet_name,
                        startcol=0, startrow=start_row - num_failed)

    def save_simu_details(self, num_failed, file_name):
        """

        Returns
        -------

        """

        "---- event data abstraction ----"
        case_id = self.case_id

        # lt track (observed in simulation and ground truth in nds)
        simu_trj_lt = self.agent_lt.observed_trajectory[:, 0:2]
        nds_trj_lt = np.array(self.lt_actual_trj[:, 0:2])
        vel_ob_vel_norm_lt = np.linalg.norm(self.agent_lt.observed_trajectory[:, 2:4], axis=1)
        vel_nds_vel_norm_lt = np.linalg.norm(self.lt_actual_trj[:, 2:4], axis=1)

        # gs track (observed in simulation and ground truth in nds)
        simu_trj_gs = self.agent_gs.observed_trajectory[:, 0:2]
        nds_trj_gs = np.array(self.gs_actual_trj[:, 0:2])
        vel_ob_vel_norm_gs = np.linalg.norm(self.agent_gs.observed_trajectory[:, 2:4], axis=1)
        vel_nds_vel_norm_gs = np.linalg.norm(self.gs_actual_trj[:, 2:4], axis=1)

        "---- sava data ----"
        # prepare data
        # simu_trj = np.concatenate((simu_trj_lt, simu_trj_gs), axis=1)
        df_simu_lt = pd.DataFrame(simu_trj_lt, columns=[str(case_id)+'-x', 'y'])
        df_simu_gs = pd.DataFrame(simu_trj_gs, columns=[str(case_id)+'-x', 'y'])
        df_nds_lt = pd.DataFrame(nds_trj_lt, columns=[str(case_id)+'-x', 'y'])
        df_nds_gs = pd.DataFrame(nds_trj_gs, columns=[str(case_id)+'-x', 'y'])

        df_simu_v_lt = pd.DataFrame(vel_ob_vel_norm_lt, columns=[str(case_id)])
        df_simu_v_gs = pd.DataFrame(vel_ob_vel_norm_gs, columns=[str(case_id)])
        df_nds_v_lt = pd.DataFrame(vel_nds_vel_norm_lt, columns=[str(case_id)])
        df_nds_v_gs = pd.DataFrame(vel_nds_vel_norm_gs, columns=[str(case_id)])

        book = load_workbook(file_name)
        # # write data
        with pd.ExcelWriter(file_name, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}

            df_simu_lt.to_excel(writer, header=True, index=False,
                                sheet_name='simulation LT Trajectory',
                                startcol=(case_id - num_failed) * 2)

            df_simu_gs.to_excel(writer, header=True, index=False,
                                sheet_name='simulation GS Trajectory',
                                startcol=(case_id - num_failed) * 2)

            df_nds_lt.to_excel(writer, header=True, index=False,
                               sheet_name='Ground truth LT Trajectory',
                               startcol=(case_id - num_failed) * 2)

            df_nds_gs.to_excel(writer, header=True, index=False,
                               sheet_name='Ground truth GS Trajectory',
                               startcol=(case_id - num_failed) * 2)

            df_nds_v_lt.to_excel(writer, header=True, index=False,
                                 sheet_name='Ground truth LT velocity',
                                 startcol=case_id - num_failed)

            df_nds_v_gs.to_excel(writer, header=True, index=False,
                                 sheet_name='Ground truth GS velocity',
                                 startcol=case_id - num_failed)

            df_simu_v_lt.to_excel(writer, header=True, index=False,
                                  sheet_name='Simulation LT velocity',
                                  startcol=case_id - num_failed)

            df_simu_v_gs.to_excel(writer, header=True, index=False,
                                  sheet_name='Simulation GS velocity',
                                  startcol=case_id - num_failed)

            writer.save()

    def save_conv_meta(self, trajectory_collection, ipv_estimation_collection, file_name, draw=False):
        """

        Returns
        -------

        """

        "---- event data abstraction ----"
        # interaction strength: distance between self-opt and interactive plan for quantifying interaction strength

        strength_lt = []  # actual: distance between opt and nds
        ave_strength_lt = []  # simulated average: mean distance between self-opt and simulated cases
        std_strength_lt = []
        min_self_strength_lt = []  # simulated minimum: minimal distance between self-opt and simulated cases
        min_self_strength_ipv_lt = []  # self-ipv that minimize the self interaction strength

        strength_gs = []
        ave_strength_gs = []
        std_strength_gs = []
        min_self_strength_gs = []
        min_self_strength_ipv_gs = []

        strength_overall = []  # sum of the interaction strength of lt and gs
        min_overall_strength_ipv_lt = []  # the ipv selection of lt that minimize the overall strength
        min_overall_strength_ipv_gs = []

        min_overall_strength_lt = []  # interaction strength minimized by lt
        ave_overall_strength_lt = []  # average interaction strength in cases from lt's perspective
        std_overall_strength_lt = []

        min_overall_strength_gs = []
        ave_overall_strength_gs = []
        std_overall_strength_gs = []

        ave_semantic_res = []
        semantic_res_std = []

        sensi_r2u_actual = []
        sensi_r2u_simu_ave_lt = []
        sensi_r2u_simu_ave_gs = []
        sensi_r2u_simu_min_lt = []
        sensi_r2u_simu_min_gs = []

        min_sensitivity_ipv_lt = []
        min_sensitivity_ipv_gs = []

        ipv_lt = []  # estimated ipv expression
        ipv_gs = []

        ipv_lt_coll = []
        ipv_gs_coll = []
        for lt_ipv in {-2, -1, -0.5, 0, 0.5, 1, 2}:
            for gs_ipv in {-2, -1, -0.5, 0, 0.5, 1, 2}:
                if 3 > lt_ipv + gs_ipv >= -0.5:
                    ipv_lt_coll.append(lt_ipv)
                    ipv_gs_coll.append(gs_ipv)

        for t in range(len(trajectory_collection)):
            trj_package = trajectory_collection[t]

            nds_trj_lt = np.array(self.lt_actual_trj[t:, 0:2])
            nds_trj_gs = np.array(self.gs_actual_trj[t:, 0:2])
            compare_range = range(min(19, np.size(nds_trj_lt, 0)))
            nds_trj_lt = nds_trj_lt[compare_range, :]
            nds_trj_gs = nds_trj_gs[compare_range, :]

            dyna_mat = [2 * i + 1 for i in compare_range]
            dyna_mat = np.array([dyna_mat, dyna_mat])
            dyna_mat = dyna_mat.T

            # optimal plan in non-interactive simulation
            self_opt_lt = trj_package['non-interactive']['lt'][compare_range, :]
            self_opt_gs = trj_package['non-interactive']['gs'][compare_range, :]

            if draw:
                plt.figure(1)
                plt.plot(self_opt_lt[:, 0], self_opt_lt[:, 1], color='green')
                plt.plot(self_opt_gs[:, 0], self_opt_gs[:, 1], color='green')

            self_plan_dev_lt = []
            inter_plan_dev_lt = []
            self_plan_dev_gs = []
            inter_plan_dev_gs = []
            sensi_r2u_lt = []
            sensi_r2u_gs = []
            semantic_res = []
            likeness = []
            for task_id in range(len(trj_package) - 1):
                # lt track (observed in simulation)
                self_plan_lt = trj_package['task' + str(task_id)]['lt-self'][compare_range, :]
                inter_plan_lt = trj_package['task' + str(task_id)]['lt-inter'][compare_range, :]

                # gs track (observed in simulation and ground truth in nds)
                self_plan_gs = trj_package['task' + str(task_id)]['gs-self'][compare_range, :]
                inter_plan_gs = trj_package['task' + str(task_id)]['gs-inter'][compare_range, :]
                if draw:
                    plt.plot(self_plan_lt[:, 0], self_plan_lt[:, 1], color='blue')
                    # plt.plot(inter_plan_lt[:, 0], inter_plan_lt[:, 1], color='blue')
                    plt.plot(self_plan_gs[:, 0], self_plan_gs[:, 1], color='purple')
                    # plt.plot(inter_plan_gs[:, 0], inter_plan_gs[:, 1], color='purple')

                # data preparation for event-level ipv estimation
                # rel_dis_lt = np.linalg.norm(self_plan_lt[:6, :] - nds_trj_lt[:6, :], axis=1)
                # rel_dis_gs = np.linalg.norm(self_plan_gs[:6, :] - nds_trj_gs[:6, :], axis=1)
                #
                # rel_dis = np.concatenate((rel_dis_lt, rel_dis_gs), axis=0)
                # likeness_temp = np.power(
                #     np.prod(
                #         (1 / sigma / np.sqrt(2 * math.pi))
                #         * np.exp(- rel_dis ** 2 / (2 * sigma ** 2)))
                #     , 1 / np.size(rel_dis, 0))
                # if likeness_temp < 0:
                #     likeness_temp = 0
                #
                # likeness.append(likeness_temp)

                "interaction strength indicated by plan deviation in each task"
                self_plan_dev_lt.append(np.mean(np.linalg.norm(self_opt_lt - self_plan_lt, axis=1)))
                inter_plan_dev_lt.append(np.mean(np.linalg.norm(self_opt_gs - inter_plan_lt, axis=1)))

                self_plan_dev_gs.append(np.mean(np.linalg.norm(self_opt_gs - self_plan_gs, axis=1)))
                inter_plan_dev_gs.append(np.mean(np.linalg.norm(self_opt_lt - inter_plan_gs, axis=1)))

                "interaction strength indicated by reward sensitivity to next action"
                plan_vec_lt = self_plan_lt - inter_plan_lt
                plan_dis_lt = np.linalg.norm(plan_vec_lt, axis=1)
                plan_vec_lt[np.where(plan_dis_lt > INTERACTION_DIS), :] = 0
                sensi_r2p_lt = plan_vec_lt / np.linalg.norm(plan_dis_lt) * 0.5 * 0.12 ** 2
                sensi_r2u_lt.append(np.sum(abs(dyna_mat * sensi_r2p_lt)))

                plan_vec_gs = self_plan_gs - inter_plan_gs
                plan_dis_gs = np.linalg.norm(plan_vec_gs, axis=1)
                plan_vec_gs[np.where(plan_dis_gs > INTERACTION_DIS), :] = 0
                sensi_r2p_gs = plan_vec_gs / np.linalg.norm(plan_dis_gs) * 0.5 * 0.12 ** 2
                sensi_r2u_gs.append(np.sum(abs(dyna_mat * sensi_r2p_gs)))

                "closest time point"
                dis_lt = np.linalg.norm(np.array(self_plan_lt) - np.array(inter_plan_lt), axis=1)
                min_dis_index_lt = np.where(min(dis_lt) == dis_lt)
                dis_gs = np.linalg.norm(np.array(self_plan_gs) - np.array(inter_plan_gs), axis=1)
                min_dis_index_gs = np.where(min(dis_gs) == dis_gs)

                "yield or rush?"
                if self_plan_lt[min_dis_index_lt[0][0], 0] > inter_plan_lt[min_dis_index_lt[0][0], 0]:
                    semantic_res.append(1)  # left-turn agent rushed
                else:
                    semantic_res.append(0)

                if self_plan_gs[min_dis_index_gs[0][0], 0] < inter_plan_gs[min_dis_index_gs[0][0], 0]:
                    semantic_res.append(1)  # left-turn agent rushed
                else:
                    semantic_res.append(0)
            if draw:
                plt.show()

            "event-level ipv estimation"
            # likeness = likeness / (sum(likeness))
            # ipv_lt.append(np.dot(likeness, ipv_lt_coll))
            # ipv_gs.append(np.dot(likeness, ipv_gs_coll))

            "actual interaction strength indicated by plan deviation"
            strength_lt.append(
                np.mean(np.linalg.norm(self_opt_lt[compare_range, :] - nds_trj_lt[compare_range, :], axis=1)))
            strength_gs.append(
                np.mean(np.linalg.norm(self_opt_gs[compare_range, :] - nds_trj_gs[compare_range, :], axis=1)))
            strength_overall.append(
                np.mean(np.linalg.norm(self_opt_lt[compare_range, :] - nds_trj_lt[compare_range, :], axis=1))
                + np.mean(np.linalg.norm(self_opt_gs[compare_range, :] - nds_trj_gs[compare_range, :], axis=1)))

            "simulated interaction strength of single agent"
            ave_strength_lt.append(np.mean(self_plan_dev_lt))
            std_strength_lt.append(np.std(self_plan_dev_lt))
            ave_strength_gs.append(np.mean(self_plan_dev_gs))
            std_strength_gs.append(np.std(self_plan_dev_gs))

            "find the ipv that minimize the ## overall ## interaction strength"
            overall_plan_dev_lt = np.array(self_plan_dev_lt) + np.array(inter_plan_dev_lt)
            min_overall_strength_id_lt = np.where(min(overall_plan_dev_lt) == overall_plan_dev_lt)
            min_overall_strength_ipv_lt.append(ipv_lt_coll[min_overall_strength_id_lt[0][0]] * math.pi / 8)
            min_overall_strength_lt.append(min(overall_plan_dev_lt))
            ave_overall_strength_lt.append(np.mean(overall_plan_dev_lt))
            std_overall_strength_lt.append(np.std(overall_plan_dev_lt))

            overall_plan_dev_gs = np.array(self_plan_dev_gs) + np.array(inter_plan_dev_gs)
            min_overall_strength_id_gs = np.where(min(overall_plan_dev_gs) == overall_plan_dev_gs)
            min_overall_strength_ipv_gs.append(ipv_gs_coll[min_overall_strength_id_gs[0][0]] * math.pi / 8)
            min_overall_strength_gs.append(min(overall_plan_dev_gs))
            ave_overall_strength_gs.append(np.mean(overall_plan_dev_gs))
            std_overall_strength_gs.append(np.std(overall_plan_dev_gs))

            "find the ipv that minimize the ## self ## interaction strength"
            min_self_strength_id_lt = np.where(min(self_plan_dev_lt) == self_plan_dev_lt)
            min_self_strength_ipv_lt.append(ipv_lt_coll[min_self_strength_id_lt[0][0]] * math.pi / 8)
            min_self_strength_lt.append(min(self_plan_dev_lt))

            min_self_strength_id_gs = np.where(min(self_plan_dev_gs) == self_plan_dev_gs)
            min_self_strength_ipv_gs.append(ipv_gs_coll[min_self_strength_id_gs[0][0]] * math.pi / 8)
            min_self_strength_gs.append(min(self_plan_dev_gs))

            "find semantic results"
            ave_semantic_res.append(np.mean(semantic_res))
            semantic_res_std.append(np.std(semantic_res))

            "interaction strength with sensitivity analysis"
            # actual
            plan_vec_act = nds_trj_lt - nds_trj_gs
            plan_dis_act = np.linalg.norm(plan_vec_act, axis=1)
            # print('min dis: ', min(plan_dis_act))
            plan_vec_act[np.where(plan_dis_act > INTERACTION_DIS), :] = 0
            sensi_r2p_actual = plan_vec_act / np.linalg.norm(plan_dis_act) * 0.5 * 0.12 ** 2
            sensi_r2u_actual.append(np.sum(abs(dyna_mat * sensi_r2p_actual)))

            # simulated average
            sensi_r2u_simu_ave_lt.append(np.mean(sensi_r2u_lt))
            sensi_r2u_simu_ave_gs.append(np.mean(sensi_r2u_gs))

            # simulated minimum
            sensi_r2u_simu_min_lt.append(min(sensi_r2u_lt))
            sensi_r2u_simu_min_gs.append(min(sensi_r2u_gs))

            # ipv that minimize the overall interaction strength
            min_sensi_id_lt = np.where(sensi_r2u_lt == min(sensi_r2u_lt))
            min_sensi_id_gs = np.where(sensi_r2u_gs == min(sensi_r2u_gs))
            min_sensitivity_ipv_lt.append(ipv_lt_coll[min_sensi_id_lt[0][0]] * math.pi / 8)
            min_sensitivity_ipv_gs.append(ipv_gs_coll[min_sensi_id_gs[0][0]] * math.pi / 8)

        "---- sava data ----"
        # prepare data
        df = pd.DataFrame([[strength_lt[i], min_self_strength_lt[i], ave_strength_lt[i], std_strength_lt[i],
                            strength_gs[i], min_self_strength_gs[i], ave_strength_gs[i], std_strength_gs[i],
                            strength_overall[i],
                            min_overall_strength_lt[i], ave_overall_strength_lt[i], std_overall_strength_lt[i],
                            min_overall_strength_gs[i], ave_overall_strength_gs[i], std_overall_strength_gs[i],
                            min_overall_strength_ipv_lt[i], min_self_strength_ipv_lt[i],
                            ipv_estimation_collection[i][0],
                            min_overall_strength_ipv_gs[i], min_self_strength_ipv_gs[i],
                            ipv_estimation_collection[i][1],
                            ave_semantic_res[i], semantic_res_std[i],
                            sensi_r2u_actual[i], sensi_r2u_simu_ave_lt[i], sensi_r2u_simu_min_lt[i],
                            sensi_r2u_simu_ave_gs[i], sensi_r2u_simu_min_gs[i],
                            ] for i in range(len(ave_strength_lt))],
                          columns=['lt-self-str-actual', '-min', '-ave', '-std',
                                   'gs-self-str-actual', '-min', '-ave', '-std',
                                   'overall-str-actual',
                                   'min_OA_str from lt', 'ave_OA_str from lt', 'std_OA_str from lt',
                                   'min_OA_str from gs', 'ave_OA_str from gs', 'std_OA_str from gs',
                                   'ipv_lt leads min_OA_str', 'ipv lt leads min_self_str', 'lt ipv',
                                   'ipv_gs leads min_OA_str', 'ipv gs leads min_self_str', 'gs ipv',
                                   'ave semantic res', 'semantic res std',
                                   'sensi_actual', 'ave sensi lt', 'min sensi lt',
                                   'ave sensi gs', 'min sensi gs',
                                   ])
        # df = pd.DataFrame([[ipv_estimation_collection[i][0], min_sensitivity_ipv_lt[i],
        #                     ipv_estimation_collection[i][1], min_sensitivity_ipv_gs[i],
        #                     ave_semantic_res[i], semantic_res_std[i],
        #                     sensi_r2u_actual[i], sensi_r2u_simu_ave_lt[i], sensi_r2u_simu_min_lt[i],
        #                     sensi_r2u_simu_ave_gs[i], sensi_r2u_simu_min_gs[i],
        #                     ] for i in range(len(ave_semantic_res))],
        #                   columns=['lt ipv', 'min sensi lt ipv',
        #                            'gs ipv', 'min sensi gs ipv',
        #                            'ave semantic res', 'semantic res std',
        #                            'sensi_actual', 'ave sensi lt', 'min sensi lt',
        #                            'ave sensi gs', 'min sensi gs',
        #                            ])

        # write data
        # book = load_workbook(file_name)
        with pd.ExcelWriter(file_name, mode='a', if_sheet_exists="overlay", engine="openpyxl") as writer:
            # writer.book = book
            df.to_excel(writer, header=True, index=False,
                        sheet_name=str(self.case_id),
                        startcol=0)

            # for column in df:
            #     column_length = max(df[column].astype(str).map(len).max(), len(column))
            #     col_idx = df.columns.get_loc(column)
            #     writer.sheets[str(self.case_id)].set_column(col_idx, col_idx, 1.2 * column_length)

            # writer.save()
            # writer.close()


def get_semantic_result(track_1, track_2, case_type='simu_left_turn'):
    """
        Identify semantic interaction results after simulation:
        1. crashed or not (not critical judgement)
        2. the left-turn vehicle yield or not
        """

    pos_delta = track_2 - track_1
    dis_delta = np.linalg.norm(pos_delta[:, 0:2], axis=1)

    if min(dis_delta) < 1:
        semantic_result = 'crashed'
        # print('interaction is crashed. \n')
    elif case_type == 'nds':
        pos_y_larger = pos_delta[pos_delta[:, 1] > 0]
        if np.size(pos_y_larger, 0):

            "whether the LT vehicle yield"
            pos_x_larger = pos_y_larger[pos_y_larger[:, 0] > 0]
            yield_points = np.size(pos_x_larger, 0)
            if yield_points:
                semantic_result = 'yield'

                "where the interaction finish"
                ind_coll = np.where(pos_x_larger[0, 0] == pos_delta[:, 0])
                ind = ind_coll[0] - 1
                ending_point = {'lt': track_1[ind, :],
                                'gs': track_2[ind, :]}

                # print('LT vehicle yielded. \n')

            else:
                semantic_result = 'rush'
                # print('LT vehicle rushed. \n')
        else:
            pos_x_smaller = pos_delta[pos_delta[:, 0] < -1]
            if np.size(pos_x_smaller, 0):
                semantic_result = 'rush'
                # print('LT vehicle rushed. \n')
            else:
                semantic_result = 'unfinished'
                # print('interaction is not finished. \n')

    elif case_type == 'simu_left_turn':
        pos_x_smaller = pos_delta[pos_delta[:, 0] < 0]
        if np.size(pos_x_smaller, 0):

            "whether the LT vehicle yield"
            pos_y_larger = pos_x_smaller[pos_x_smaller[:, 1] > 0]
            yield_points = np.size(pos_y_larger, 0)
            if yield_points:
                semantic_result = 'yield'

                "where the interaction finish"
                ind_coll = np.where(pos_y_larger[0, 0] == pos_delta[:, 0])
                ind = ind_coll[0] - 1
                ending_point = {'lt': track_1[ind, :],
                                'gs': track_2[ind, :]}

                # print('LT vehicle yielded. \n')

            else:
                semantic_result = 'rush'
                # print('LT vehicle rushed. \n')
        else:
            pos_y_smaller = pos_delta[pos_delta[:, 1] < 0]
            if np.size(pos_y_smaller, 0):
                semantic_result = 'rush'
                # print('LT vehicle rushed. \n')
            else:
                semantic_result = 'unfinished'
                # print('interaction is not finished. \n')
    else:  # case_type == 'simu_ramp':
        if track_1[-1, 0] > track_2[-1, 0]:
            semantic_result = 'main line first'
        else:
            semantic_result = 'ramp first'

    return semantic_result


def run_interaction(case_id, task_id, t, lt_ipv, gs_ipv, returns, con_type='linear-gt'):
    simu = Simulator(case_id=case_id)
    simu.sim_type = 'nds'
    controller_type_lt = con_type
    controller_type_gs = con_type
    simu.read_nds_scenario(controller_type_lt, controller_type_gs)

    simu.simu_time = t
    tag = 'conv analysis-case' + str(case_id) + '-t' + str(t) + '-task' + str(task_id)
    simu_scenario = simu.read_nds_scenario(controller_type_lt, controller_type_gs, t=t)

    simu.initialize(simu_scenario, tag)

    simu.agent_gs.target = 'gs_nds'
    simu.agent_gs.estimated_inter_agent[0].target = 'lt_nds'
    simu.agent_lt.target = 'lt_nds'
    simu.agent_lt.estimated_inter_agent[0].target = 'gs_nds'

    simu.agent_gs.ipv = gs_ipv * math.pi / 8
    simu.agent_gs.estimated_inter_agent[0].ipv = lt_ipv * math.pi / 8
    simu.agent_lt.ipv = lt_ipv * math.pi / 8
    simu.agent_lt.estimated_inter_agent[0].ipv = gs_ipv * math.pi / 8

    simu.interact(simu_step=1)
    # simu.visualize_single_step(file_path='../outputs/5_gt_interaction/figures/convergence analysis-'
    #                                      + str(case_id) + '/')

    # print('simulation finished: task id - ' + str(task_id))

    returns['task' + str(task_id)] = {'lt-self': simu.agent_lt.trj_solution[:, 0:2],
                                      'lt-inter': simu.agent_lt.estimated_inter_agent[0].trj_solution[:, 0:2],
                                      'gs-self': simu.agent_gs.trj_solution[:, 0:2],
                                      'gs-inter': simu.agent_gs.estimated_inter_agent[0].trj_solution[:, 0:2],
                                      'ipv': [lt_ipv, gs_ipv]}

    return returns


def run_interaction_multi(case_id, task_id, t, lt_ipv, gs_ipv, gs_id, returns):
    simu = Simulator(case_id=case_id)
    simu.sim_type = 'nds'
    controller_type_lt = 'linear-gt'
    controller_type_gs = 'linear-gt'
    simu.read_nds_scenario(controller_type_lt, controller_type_gs)

    simu.simu_time = t
    simu.gs_id = gs_id
    tag = 'conv analysis-case' + str(case_id) + '-t' + str(t) + '-task' + str(task_id)
    simu_scenario = simu.read_nds_scenario_multi(controller_type_lt, controller_type_gs, t=t)

    simu.initialize(simu_scenario, tag)

    simu.agent_gs.target = 'gs_nds'
    simu.agent_gs.estimated_inter_agent.target = 'lt_nds'
    simu.agent_lt.target = 'lt_nds'
    simu.agent_lt.estimated_inter_agent.target = 'gs_nds'

    simu.agent_gs.ipv = gs_ipv * math.pi / 8
    simu.agent_gs.estimated_inter_agent.ipv = lt_ipv * math.pi / 8
    simu.agent_lt.ipv = lt_ipv * math.pi / 8
    simu.agent_lt.estimated_inter_agent.ipv = gs_ipv * math.pi / 8

    simu.interact(simu_step=1)
    # simu.visualize_single_step(file_path='../outputs/5_gt_interaction/figures/convergence analysis-'
    #                                      + str(case_id) + '/')

    # print('simulation finished: task id - ' + str(task_id))

    returns['task' + str(task_id)] = {'lt-self': simu.agent_lt.trj_solution[:, 0:2],
                                      'lt-inter': simu.agent_lt.estimated_inter_agent.trj_solution[:, 0:2],
                                      'gs-self': simu.agent_gs.trj_solution[:, 0:2],
                                      'gs-inter': simu.agent_gs.estimated_inter_agent.trj_solution[:, 0:2]}

    return returns


def main_simulate_t_intersection():
    """
    === main for simulating unprotected left-turning ===
    1. Set initial motion state before the simulation
    2. Change controller type by manually setting controller_type_xx as:
        * 'gt' is game-theoretic planner work by solving IBR
        * 'linear-gt' is a linear game-theoretic planner work by solving IBR
        * 'opt' is the optimal controller work by solving single optimization
    3. a simple lattice planner is applied as the algorithm under test
    """
    case_id = 0
    param_v = 1
    role_under_test = None
    bg_ipv = 1 * math.pi / 8  # ipv of background vehicle
    tag = 'test'  # tag for data saving

    '---- set initial state of the left-turn vehicle ----'
    init_position_lt = [8.5, -8.5]
    init_velocity_lt = [1, 2] * param_v
    init_heading_lt = math.pi / 3
    ipv_lt = bg_ipv
    controller_type_lt = 'gt'
    if role_under_test == 'lt':
        controller_type_lt = 'lattice'

    '---- set initial state of the go-straight vehicle ----'
    init_position_gs = [21, -2]
    init_velocity_gs = [-3, 0] * param_v
    init_heading_gs = math.pi
    ipv_gs = bg_ipv
    controller_type_gs = 'linear-gt'
    if role_under_test == 'gs':
        controller_type_gs = 'lattice'

    '---- generate scenario ----'
    simu_scenario = Scenario([init_position_lt, init_position_gs],
                             [init_velocity_lt, init_velocity_gs],
                             [init_heading_lt, init_heading_gs],
                             [ipv_lt, ipv_gs],
                             [controller_type_lt, controller_type_gs])

    simu = Simulator(case_id=case_id)
    simu.sim_type = 'simu_left_turn'
    simu.initialize(simu_scenario, tag)  # initialize the agents in the simulator

    time1 = time.perf_counter()
    simu.interact(simu_step=35, iter_limit=5, make_video=True)
    time2 = time.perf_counter()
    print('time consumption: ', time2 - time1)

    simu.semantic_result = get_semantic_result(simu.agent_lt.observed_trajectory, simu.agent_gs.observed_trajectory,
                                               simu.sim_type)  # find semantic result (yield or rush)

    # print final trajectory at given path
    simu.visualize_final_results(file_path='../outputs/5_gt_interaction/figures/')

    # simu.save_test_meta(param_v=param_v, ipv=bg_ipv, file_name='outputs/test_meta_data.xlsx',
    #                     sheet_name='left-turn test' + role_under_test)


def main_simulate_ramp_merge():
    """
    === main for simulating ramp merging ===
    1. Set initial motion state before the simulation
    2. Change controller type by manually setting controller_type_xx as:
        * 'gt' is game-theoretic planner work by solving IBR
        * 'linear-gt' is a linear game-theoretic planner work by solving IBR
        * 'opt' is the optimal controller work by solving single non-linear optimization
    3. a simple lattice planner is applied as the algorithm under test
    """
    case_id = 0

    role_under_test = 'ir'  # ir: in-ramp, ml: main line
    # for param in range(0, 20):
    for param in {13}:
        param_v = param / 10
        # for param_ipv in range(-7, 8):
        for param_ipv in {0}:

            print('test-' + role_under_test + '- case: ' + str(case_id))

            bg_ipv = param_ipv / 2 * math.pi / 8
            tag = 'test'  # tag for data saving

            '---- set initial state of the main line vehicle ----'
            init_position_ml = [-26, -2]
            init_velocity_ml = np.array([5, 0])
            init_heading_ml = 0
            ipv_ml = bg_ipv
            controller_type_ml = 'linear-gt'
            if role_under_test == 'ml':
                controller_type_ml = 'lattice'
            else:
                init_velocity_ml = init_velocity_ml * param_v

            '---- set initial state of the merging vehicle ----'
            init_position_ir = [-25, -7.5]
            init_velocity_ir = np.array([5, 1])
            init_heading_ir = 0
            ipv_ir = bg_ipv
            controller_type_ir = 'linear-gt'
            if role_under_test == 'ir':
                controller_type_ir = 'lattice'
            else:
                init_velocity_ir = init_velocity_ir * param_v

            '---- generate scenario ----'
            simu_scenario = Scenario([init_position_ml, init_position_ir],
                                     [init_velocity_ml, init_velocity_ir],
                                     [init_heading_ml, init_heading_ir],
                                     [ipv_ml, ipv_ir],
                                     [controller_type_ml, controller_type_ir], sce_type='ramp merge')

            simu = Simulator(case_id=case_id)
            simu.sim_type = 'simu_ramp'
            simu.initialize(simu_scenario, tag)

            time1 = time.perf_counter()
            simu.interact(simu_step=30, iter_limit=5, make_video=False)
            time2 = time.perf_counter()
            print('time consumption: ', time2 - time1)

            simu.semantic_result = get_semantic_result(simu.agent_lt.observed_trajectory,
                                                       simu.agent_gs.observed_trajectory, simu.sim_type)

            simu.visualize_final_results(
                file_path='../outputs/5_gt_interaction/figures/')  # print final trajectory at given path

            # simu.save_test_meta(param_v=param_v, ipv=bg_ipv,
            #                     file_name='outputs/test_meta_data-' + role_under_test + '.xlsx',
            #                     sheet_name='ramp test-' + role_under_test)
            case_id += 1


def main_simulate_nds():
    """
       === main for simulating unprotected left-turning scenarios in Jianhe-Xianxia dataset ===
       1. Set case_id to get initial scenarios state of a single case
       2. Change controller type by manually setting controller_type_xx as:
           * 'gt' is the game-theoretic planner work by solving IBR process
           * 'opt' is the optimal controller work by solving single optimization
           * 'linear-gt' is a linear game-theoretic planner work by solving IBR
           * 'idm'
           * 'replay' *** Set dt as 0.12 (in agent.py) before simulation ***
       """

    model_type = 'opt'
    target = 'simu'

    data_path = '../outputs/5_gt_interaction/data_records/' \
                + target + '/' + model_type + '/'
    if not os.path.exists(data_path):
        os.makedirs(data_path)

    start_time = strftime("%Y-%m-%d-%H", gmtime())
    file_name = data_path + 'simulation_detail-' + start_time + '.xlsx'
    workbook = xlsxwriter.Workbook(file_name)
    workbook.close()

    num_failed = 0

    for case_id in range(130):
        # for case_id in {51}:

        if case_id in {39, 45, 78, 93, 99}:  # interactions finished at the beginning
            num_failed += 1
            continue
        elif case_id in {12, 13, 24, 26, 28, 31, 32, 33, 37, 38, 46, 47, 48, 52, 56, 59, 65, 66, 69, 77, 82, 83, 84, 90,
                         91, 92, 94, 96, 97, 98, 100}:  # no path-crossing event
            num_failed += 1
        else:
            "1. tag for simulation"
            tag = target + '-' + model_type + str(case_id)

            "2. tag for testing"
            # tag = model_type + '-test-lattice' + str(case_id)

            print('start case:' + tag)

            simu = Simulator(case_id=case_id)
            simu.sim_type = 'nds'
            controller_type_lt = model_type
            controller_type_gs = model_type
            simu_scenario = simu.read_nds_scenario(controller_type_lt, controller_type_gs)
            if simu_scenario:
                simu.initialize(simu_scenario, tag)

                simu.agent_gs.target = 'gs_nds'
                for _, inter_agent in enumerate(simu.agent_gs.estimated_inter_agent):
                    inter_agent.target = 'lt_nds'

                simu.agent_lt.target = 'lt_nds'
                for _, inter_agent in enumerate(simu.agent_lt.estimated_inter_agent):
                    inter_agent.target = 'gs_nds'

                # simu.agent_gs.estimated_inter_agent[0].ipv = simu.agent_lt.ipv
                # simu.agent_lt.estimated_inter_agent[0].ipv = simu.agent_gs.ipv

                try:
                    fig_path = '../outputs/5_gt_interaction/figures/' + target + '/' \
                               + model_type + '-case-' + str(simu.case_id) + '/'
                    if not os.path.exists(fig_path):
                        os.makedirs(fig_path)

                    time1 = time.perf_counter()
                    simu.interact(simu_step=int(simu.case_len),
                                  make_video=False,
                                  break_when_finish=False,
                                  file_path=fig_path)
                    time2 = time.perf_counter()
                    print('time consumption: ', time2 - time1)

                    # ----print final trajectory at given path
                    # simu.visualize_final_results(file_path=fig_path)

                    # ----get semantic interaction result
                    # simu.semantic_result = get_semantic_result(simu.agent_lt.observed_trajectory[:, 0:2],
                    #                                            simu.agent_gs.observed_trajectory[:, 0:2],
                    #                                            case_type='nds')

                    # ----save meta info of each interaction simulation
                    # simu.save_simu_meta(num_failed, file_name=fig_path + 'simulation_meta_data.xlsx',
                    #                     sheet_name=model_type + ' simulation')

                    # ----save detailed trajectory of each simulation

                    simu.save_simu_details(num_failed, file_name=file_name)

                except IndexError:
                    print('# ====Failed:' + tag + '==== #')
                    num_failed += 1
                    continue
            else:
                num_failed += 1


def main_analyze_interaction_strength_v1():
    """
    main for analyzing interaction strength and convergence in nds cases where LT car interact with a single GS car
    Note: in this version, we solve game at each time step and calculate info with data in a fixed future horizon
    """

    bg_type = 'gt'
    num_failed = 0
    file_name = '../outputs/5_gt_interaction/outputs/conv_meta_data20230112.xlsx'
    if not os.path.exists(file_name):
        workbook = xlsxwriter.Workbook(file_name)
        workbook.add_worksheet()
        workbook.close()

    proc_bar = tqdm(range(0, 130))
    # proc_bar = tqdm({50})
    for case_id in proc_bar:

        if case_id in {39, 45, 78, 93, 99}:  # interactions finished at the beginning
            num_failed += 1
            continue
        elif case_id in {12, 13, 24, 26, 28, 31, 32, 33, 37, 38, 46, 47, 48, 52, 56, 59, 65, 66, 69, 77, 82, 83, 84, 90,
                         91, 92, 94, 96, 97, 98, 100}:  # no path-crossing event
            num_failed += 1
            continue
        elif case_id in {7, 23, 53, 54, 55, 79, 112, 114, 115, 116, 129}:  # influenced by non-moter road users
            num_failed += 1
            continue
        else:

            # print('start case:' + str(case_id))

            simu = Simulator(case_id=case_id)
            simu.sim_type = 'nds'
            controller_type_lt = bg_type
            controller_type_gs = bg_type
            simu.read_nds_scenario(controller_type_lt, controller_type_gs)

            trajectory_collection = []
            ipv_estimation_collection = []

            for t in range(simu.case_len - 10):
                # for t in range(1):

                proc_bar.set_postfix({"processing": f"{t}"})
                simu.simu_time = t
                tag = 'conv-analysis-case' + str(case_id) + '-t' + str(t)
                simu_scenario = simu.read_nds_scenario(controller_type_lt, controller_type_gs, t=t)
                traj_coll_temp = {}

                if simu_scenario:
                    simu.initialize(simu_scenario, tag)

                    simu.agent_gs.target = 'gs_nds'
                    simu.agent_gs.estimated_inter_agent[0].target = 'lt_nds'
                    simu.agent_lt.target = 'lt_nds'
                    simu.agent_lt.estimated_inter_agent[0].target = 'gs_nds'

                    # save origin for ipv estimation
                    agent_lt_temp = copy.deepcopy(simu.agent_lt)
                    agent_gs_temp = copy.deepcopy(simu.agent_gs)

                    "individual-level ipv estimation"
                    nds_trj_lt = np.array(simu.lt_actual_trj[t:, 0:2])
                    nds_trj_gs = np.array(simu.gs_actual_trj[t:, 0:2])
                    compare_range = range(min(10, np.size(nds_trj_lt, 0)))
                    nds_trj_lt = nds_trj_lt[compare_range, :]
                    nds_trj_gs = nds_trj_gs[compare_range, :]

                    agent_lt_temp.estimate_self_ipv(nds_trj_lt, nds_trj_gs)
                    agent_gs_temp.estimate_self_ipv(nds_trj_gs, nds_trj_lt)

                    ipv_estimation_collection.append([agent_lt_temp.ipv, agent_gs_temp.ipv])

                    # worker pairs for generating trajectory under different ipv combinations
                    tasks = []
                    manager = Manager()
                    traj_coll_temp = manager.dict()
                    task_id = 0
                    for lt_ipv in {-2, -1, -0.5, 0, 0.5, 1, 2}:
                        for gs_ipv in {-2, -1, -0.5, 0, 0.5, 1, 2}:
                            if 3 > lt_ipv + gs_ipv >= -0.5:
                                task = Process(target=run_interaction,
                                               args=(case_id, task_id, t, lt_ipv, gs_ipv, traj_coll_temp, bg_type))
                                task.start()
                                tasks.append(task)
                                task_id += 1

                    for task in tasks:
                        task.join()

                    # generate selfish plan
                    simu.interact(simu_step=1, interactive=False)
                    simu.tag += '-self-opt'
                    # simu.visualize_single_step(file_path='./figures/convergence analysis-' + str(case_id) + '/')
                    traj_coll_temp['non-interactive'] = {'lt': simu.agent_lt.trj_solution[:, 0:2],
                                                         'gs': simu.agent_gs.trj_solution[:, 0:2]}

                trajectory_collection.append(traj_coll_temp)
                # simu.visualize_multi_interaction(trajectory_collection, t)

            simu.save_conv_meta(trajectory_collection, ipv_estimation_collection,
                                file_name=file_name,
                                draw=False)


def main_analyze_multi_interaction_strength_v1():
    """
    main for analyzing interaction strength and convergence in a case where LT vehicle interact with several GS vehicles
    """

    bg_type = 'linear-gt'
    num_failed = 0
    file_name = '../outputs/5_gt_interaction/outputs/conv_meta_data20230112.xlsx'
    if not os.path.exists(file_name):
        workbook = xlsxwriter.Workbook(file_name)
        workbook.add_worksheet()
        workbook.close()

    # for case_id in range(0, 51):
    for case_id in {51}:

        if case_id in {39, 45, 78, 93, 99}:  # interactions finished at the beginning
            num_failed += 1
            continue
        elif case_id in {12, 13, 24, 26, 28, 31, 32, 33, 37, 38, 46, 47, 48, 52, 56, 59, 65, 66, 69, 77, 82, 83, 84, 90,
                         91, 92, 94, 96, 97, 98, 100}:  # no path-crossing event
            num_failed += 1
            continue
        else:

            print('start case:' + str(case_id))

            simu = Simulator(case_id=case_id)
            simu.sim_type = 'nds'
            controller_type_lt = bg_type
            controller_type_gs = bg_type

            simu.gs_id = 2

            simu.read_nds_scenario_multi(controller_type_lt, controller_type_gs)

            trajectory_collection = []

            for t in range(simu.case_len):
                # for t in range(1):

                print('time_step: ', t, '/', simu.case_len)

                simu.simu_time = t
                tag = 'conv-analysis-case' + str(case_id) + '-t' + str(t)
                simu_scenario = simu.read_nds_scenario_multi(controller_type_lt, controller_type_gs, t=t)
                traj_coll_temp = {}

                if simu_scenario:
                    simu.initialize(simu_scenario, tag)

                    simu.agent_gs.target = 'gs_nds'
                    simu.agent_gs.estimated_inter_agent.target = 'lt_nds'
                    simu.agent_lt.target = 'lt_nds'
                    simu.agent_lt.estimated_inter_agent.target = 'gs_nds'

                    # worker pairs for generating trajectory under different ipv combinations
                    tasks = []
                    manager = Manager()
                    traj_coll_temp = manager.dict()
                    task_id = 0
                    # for lt_ipv in {-3, -1.5, 0, 1.5, 3}:
                    #     for gs_ipv in {-3, -1.5, 0, 1.5, 3}:
                    for lt_ipv in {-0.5, 0, 1}:
                        for gs_ipv in {-0.5, 0, 1}:
                            task = Process(target=run_interaction_multi,
                                           args=(case_id, task_id, t, lt_ipv, gs_ipv, simu.gs_id, traj_coll_temp))
                            task.start()
                            tasks.append(task)
                            task_id += 1

                    for task in tasks:
                        task.join()

                    # generate selfish plan
                    simu.interact(simu_step=1, interactive=False)
                    simu.tag += '-self-opt'
                    # simu.visualize_single_step(file_path='./figures/convergence analysis-' + str(case_id) + '/')
                    traj_coll_temp['non-interactive'] = {'lt': simu.agent_lt.trj_solution[:, 0:2],
                                                         'gs': simu.agent_gs.trj_solution[:, 0:2]}

                trajectory_collection.append(traj_coll_temp)
            simu.save_conv_meta(trajectory_collection, file_name=file_name)


def main_analyze_interaction_strength_v2():
    """
    simulate a full trajectory at the beginning frame of each real world case
    compare simulation results and ground truth trajectory to find the most similar simulation case
    check if the most similar one has relative low interaction strength
    Returns
    -------

    """
    bg_type = 'linear-gt'
    num_failed = 0
    dir_name = '../outputs/5_gt_interaction/inter_str_analysis_v2/'

    file_name = dir_name + 'inter_str_meta' + strftime("%Y-%m-%d", gmtime()) + '.xlsx'
    if not os.path.exists(file_name):
        workbook = xlsxwriter.Workbook(file_name)
        workbook.add_worksheet()
        workbook.close()

    # proc_bar = tqdm(range(0, 130))
    proc_bar = tqdm({106})
    for case_id in proc_bar:

        if case_id in {39, 45, 78, 93, 99}:  # interactions finished at the beginning
            num_failed += 1
            continue
        elif case_id in {12, 13, 24, 26, 28, 31, 32, 33, 37, 38, 46, 47, 48, 52, 56, 59, 65, 66, 69, 77, 82, 83, 84, 90,
                         91, 92, 94, 96, 97, 98, 100}:  # no path-crossing event
            num_failed += 1
            continue
        elif case_id in {7, 23, 53, 54, 55, 79, 112, 114, 115, 116, 129}:  # influenced by non-motor road users
            num_failed += 1
            continue
        else:

            simu = Simulator(case_id=case_id)
            simu.sim_type = 'nds'
            controller_type_lt = bg_type
            controller_type_gs = bg_type
            simu.read_nds_scenario(controller_type_lt, controller_type_gs)

            trajectory_collection = []

            # proc_bar.set_postfix({"processing": f"{t}"})
            tag = 'inter-strength-analysis-case' + str(case_id)
            simu_scenario = simu.read_nds_scenario(controller_type_lt, controller_type_gs)

            simu.initialize(simu_scenario, tag)

            simu.agent_gs.target = 'gs_nds'
            simu.agent_gs.estimated_inter_agent[0].target = 'lt_nds'
            simu.agent_lt.target = 'lt_nds'
            simu.agent_lt.estimated_inter_agent[0].target = 'gs_nds'

            # worker pairs for generating trajectory under different ipv combinations
            tasks = []
            manager = Manager()
            traj_coll_temp = manager.dict()
            task_id = 0
            simu.ipv_list = [-0.5, 0, 0.5, 1, 1.5, 2.5, 3]
            for lt_ipv in simu.ipv_list:
                for gs_ipv in simu.ipv_list:
                    # for lt_ipv in {1}:
                    #     for gs_ipv in {1}:
                    #         if 3 > lt_ipv + gs_ipv >= -0.5:
                    task = Process(target=run_interaction,
                                   args=(case_id, task_id, 0, lt_ipv, gs_ipv, traj_coll_temp, bg_type))
                    task.start()
                    tasks.append(task)
                    task_id += 1

            for task in tasks:
                task.join()

            # generate selfish plan
            simu.interact(simu_step=1, interactive=False)
            simu.tag += '-self-opt'
            # simu.visualize_single_step(file_path='./figures/convergence analysis-' + str(case_id) + '/')
            traj_coll_temp['non-interactive'] = {'lt': simu.agent_lt.trj_solution[:, 0:2],
                                                 'gs': simu.agent_gs.trj_solution[:, 0:2]}

            trajectory_collection.append(traj_coll_temp)
            axes, estimated_ipv = simu.cal_trj_similarity(trajectory_collection, isFig=False)
            inter_strength_matrix = simu.cal_interaction_strength(trajectory_collection, ax=axes,
                                                                  file_dir=dir_name,
                                                                  isFig=False,
                                                                  isSaveFig=False)
            simu.min_inter_strength_test(estimated_ipv, inter_strength_matrix, isSave=True, file_name=file_name)
            # simu.visualize_multi_interaction(trajectory_collection, 0)

        # simu.save_conv_meta(trajectory_collection, ipv_estimation_collection,
        #                         file_name=file_name,
        #                         draw=False)


if __name__ == '__main__':
    'simulate unprotected left-turn at a T-intersection'
    # main_simulate_t_intersection()

    'simulate ramp merging'
    # main_simulate_ramp_merge()

    'simulate with nds data from Jianhe-Xianxia intersection'
    main_simulate_nds()

    'analyze interaction strength by solving game at each time step'
    # main_analyze_interaction_strength_v1()
    # main_analyze_multi_interaction_strength_v1()

    'analyze interaction strength by solving game only once at the beginning'
    # main_analyze_interaction_strength_v2()
